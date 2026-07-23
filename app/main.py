"""FastAPI application: API endpoints + static frontend."""
from __future__ import annotations

import csv
import hashlib
import io
import json
import asyncio
import logging
import os
import re
import smtplib
import threading
from contextlib import asynccontextmanager, suppress
from datetime import date, datetime, timezone
from email.message import EmailMessage
from html import escape as html_escape
from pathlib import Path
from typing import Optional
from zoneinfo import ZoneInfo

# Load local/server environment variables when present. The example file is
# intentionally not loaded at runtime because it must not contain real secrets.
from dotenv import load_dotenv
_ROOT_ENV = Path(__file__).resolve().parent.parent
load_dotenv(_ROOT_ENV / ".env", override=False)

import secrets

import yaml
from fastapi import BackgroundTasks, Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import JSONResponse
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select, text
from sqlalchemy.exc import OperationalError, SQLAlchemyError
from sqlalchemy.orm import Session
from starlette.middleware.sessions import SessionMiddleware

from app import auth, ingestion, models, schemas, workflow
from app.database import SessionLocal, get_db, init_db

STATIC_DIR = Path(__file__).resolve().parent / "static"
IMAGE_DIR = Path(__file__).resolve().parent.parent / "image"
POSTGRES_SYNC_INTERVAL_SECONDS = int(os.getenv("POSTGRES_SYNC_INTERVAL_SECONDS", "1800"))
POSTGRES_AUTO_SYNC = os.getenv("POSTGRES_AUTO_SYNC", "true").lower() not in {"0", "false", "no"}
POSTGRES_SYNC_PROJECT_NAME = os.getenv("POSTGRES_SYNC_PROJECT_NAME", "Postgres Sync")
SMTP_SERVER = "192.168.100.31"
SMTP_PORT = 25
SUCURSAL_ORIGIN_EMAIL = "admjennifer@porunpaismejor.com.mx"
FRANCHISE_ORIGIN_EMAIL = "lalbornoz@farmaciasdoctorsimi.cl"
SUCURSAL_LEGAL_TO = [
    "curibe@farmaciasdoctorsimi.cl",
    "arriendos@farmaciasdoctorsimi.cl",
    "jvasquez@farmaciasdoctorsimi.cl",
]
SUCURSAL_LEGAL_CC = [
    "mcasanova@farmaciasdoctorsimi.cl",
    "lespinoza@farmaciasdoctorsimi.cl",
    "cfolsch@farmaciasdoctorsimi.cl",
    SUCURSAL_ORIGIN_EMAIL,
]
SUCURSAL_REDUCED_TO = [
    "mbustos@farmaciasdoctorsimi.cl",
    "amarquez@farmaciasdoctorsimi.cl",
    "icruz@farmaciasdoctorsimi.cl",
    "dgonzalez@farmaciasdoctorsimi.cl",
    "mmadridf@farmaciasdoctorsimi.cl",
    "rmalave@farmaciasdoctorsimi.cl",
    "ptarsetti@farmaciasdoctorsimi.cl",
    "yarevalo@farmaciasdoctorsimi.cl",
    "efredes@farmaciasdoctorsimi.cl",
    "dbustos@farmaciasdoctorsimi.cl",
    "kcarrera@farmaciasdoctorsimi.cl",
    "kleiva@farmaciasdoctorsimi.cl",
    "lberrios@farmaciasdoctorsimi.cl",
    "bdonoso@farmaciasdoctorsimi.cl",
    "arriendos@farmaciasdoctorsimi.cl",
    "emeza@farmaciasdoctorsimi.cl",
]
SUCURSAL_REDUCED_CC = [
    "mcasanova@farmaciasdoctorsimi.cl",
    "vcifuentes@farmaciasdoctorsimi.cl",
]
logger = logging.getLogger("site_swiper")
postgres_sync_lock = threading.Lock()
SANTIAGO_TZ = ZoneInfo("America/Santiago")
JEFATURA_ADMIN_EMAIL = "jef@local"
COMMERCIAL_DIVISIONS = {"SUCURSAL", "FRANQUICIA"}
JEFATURA_GROUPS = {"SUCURSAL", "FRANQUICIA", "APERTURA"}
REQUESTER_CATEGORY_EMAILS = {
    "Sucursal": {
        "admricardo@porunpaismejor.com.mx",
        "venfelipe@porunpaismejor.com.mx",
        "ventnoe@porunpaismejor.com.mx",
        "ventluis@porunpaismejor.com.mx",
        "admalemaggi@porunpaismejor.com.mx",
        "ventmarco@porunpaismejor.com.mx",
        "ventkarba@porunpaismejor.com.mx",
        "ventgerman@porunpaismejor.com.mx",
        "ventcatalina@porunpaismejor.com.mx",
        "admroberto@porunpaismejor.com.mx",
        "ventjoaravena@porunpaismejor.com.mx",
        "admivan@porunpaismejor.com.mx",
        "vensebastian@porunpaismejor.com.mx",
        "admjennifer@porunpaismejor.com.mx",
        "ventlorena@porunpaismejor.com.mx",
    },
    "Franquicia": {
        "franfrancisco@porunpaismejor.com.mx",
        "franwalter@porunpaismejor.com.mx",
        "franarnaldo@porunpaismejor.com.mx",
        "franvgarrido@porunpaismejor.com.mx",
        "franclaudio@porunpaismejor.com.mx",
        "franbastian@porunpaismejor.com.mx",
        "franmauricio@porunpaismejor.com.mx",
        "frangabriel@porunpaismejor.com.mx",
        "franalejandro@porunpaismejor.com.mx",
        "franjosev@porunpaismejor.com.mx",
        "franmaxi@porunpaismejor.com.mx",
        "francesar@porunpaismejor.com.mx",
        "franximena@porunpaismejor.com.mx",
        "franchristian@porunpaismejor.com.mx",
        "franantonio@porunpaismejor.com.mx",
    },
    "Arriendos": {"aypcelia@porunpaismejor.com.mx"},
}


@asynccontextmanager
async def lifespan(app: FastAPI):
    try:
        init_db()  # auto-create the configured database schema on first run
    except OperationalError:
        logger.warning("Database unavailable during startup; serving cache/offline mode.")
        yield
        return
    except SQLAlchemyError:
        logger.exception("Database initialization failed during startup; serving cache/offline mode.")
        yield
        return

    # Ensure a sysadmin exists so a fresh install is usable.
    db = SessionLocal()
    try:
        auth.seed_sysadmin(db)
    except OperationalError:
        logger.warning("Database unavailable while seeding sysadmin; serving cache/offline mode.")
        yield
        return
    except SQLAlchemyError:
        logger.exception("Database error while seeding sysadmin; serving cache/offline mode.")
        yield
        return
    finally:
        db.close()
    sync_task = None
    if POSTGRES_AUTO_SYNC:
        sync_task = asyncio.create_task(_postgres_sync_loop())
        app.state.postgres_sync_task = sync_task
    try:
        yield
    finally:
        if sync_task:
            sync_task.cancel()
            with suppress(asyncio.CancelledError):
                await sync_task


app = FastAPI(title="Site Swiper", version="1.0.0", lifespan=lifespan)


@app.exception_handler(OperationalError)
async def database_operational_error_handler(request: Request, exc: OperationalError):
    logger.warning("Database unavailable for %s %s", request.method, request.url.path)
    return JSONResponse(
        status_code=503,
        content={"detail": "Base de datos no disponible. La app seguira en modo cache local."},
    )


@app.exception_handler(SQLAlchemyError)
async def database_error_handler(request: Request, exc: SQLAlchemyError):
    logger.exception("Database error for %s %s", request.method, request.url.path)
    return JSONResponse(
        status_code=503,
        content={"detail": "Base de datos no disponible. La app seguira en modo cache local."},
    )

# Signed-cookie sessions. SESSION_SECRET must be set in production; for local
# dev we fall back to a random per-process key (logs everyone out on restart).
_session_secret = os.environ.get("SESSION_SECRET")
if not _session_secret:
    _session_secret = secrets.token_urlsafe(32)
    print("WARNING: SESSION_SECRET not set â€” using an ephemeral key "
          "(sessions won't survive restart). Set SESSION_SECRET for production.")
app.add_middleware(
    SessionMiddleware,
    secret_key=_session_secret,
    https_only=False,  # set True behind HTTPS in production
    same_site="lax",
)


def _versioned_image_url(url: str | None) -> str | None:
    """Append a file mtime cache-buster for locally served image assets."""
    if not url or not url.startswith("/images/"):
        return url
    filename = url.split("?", 1)[0].removeprefix("/images/")
    image_path = IMAGE_DIR / filename
    if not image_path.exists():
        return url
    return f"{url.split('?', 1)[0]}?v={int(image_path.stat().st_mtime)}"


@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/health/db")
def health_db(db: Session = Depends(get_db)):
    try:
        db.execute(text("SELECT 1"))
    except OperationalError:
        return JSONResponse(status_code=503, content={"status": "error", "database": "unavailable"})
    except SQLAlchemyError:
        logger.exception("Database healthcheck failed")
        return JSONResponse(status_code=503, content={"status": "error", "database": "error"})
    return {"status": "ok", "database": "ok"}


# --------------------------------------------------------------------------- #
# Auth endpoints
# --------------------------------------------------------------------------- #
@app.post("/auth/login", response_model=schemas.UserOut)
def login(
    payload: schemas.LoginRequest,
    request: Request,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
):
    user = db.scalar(select(models.User).where(models.User.email == payload.email))
    if user is None or not user.active or not auth.verify_password(
        payload.password, user.password_hash
    ):
        raise HTTPException(401, "Invalid email or password")
    request.session[auth.SESSION_USER_KEY] = user.id
    request.session[auth.SESSION_USER_SNAPSHOT_KEY] = {
        "id": user.id,
        "email": user.email,
        "name": user.name,
        "role": user.role,
        "commercial_division": user.commercial_division,
        "job_title": user.job_title,
        "supervisor_emails": user.supervisor_emails,
    }
    request.session["review_session_started_at"] = datetime.now(timezone.utc).isoformat()
    if POSTGRES_AUTO_SYNC:
        background_tasks.add_task(_run_postgres_sync_once, "login")
    return user


@app.post("/auth/logout")
def logout(request: Request):
    request.session.pop(auth.SESSION_USER_KEY, None)
    return {"ok": True}


@app.get("/me", response_model=schemas.UserOut)
def me(user: models.User = Depends(auth.get_current_user)):
    return user


@app.get("/sync/version")
def sync_version(
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.get_current_user),
):
    latest_review_id = db.scalar(select(func.max(models.Review.id))) or 0
    candidate_rows = db.execute(
        select(
            models.LocationCandidate.id,
            models.LocationCandidate.status,
            models.LocationCandidate.workflow_group,
            models.LocationCandidate.last_action,
            models.LocationCandidate.last_action_at,
        ).order_by(models.LocationCandidate.id)
    ).all()
    candidate_state = "|".join(
        ":".join(
            (
                str(candidate_id),
                status or "",
                workflow_group or "",
                last_action or "",
                last_action_at.isoformat() if last_action_at else "",
            )
        )
        for candidate_id, status, workflow_group, last_action, last_action_at in candidate_rows
    )
    candidate_digest = hashlib.sha256(candidate_state.encode("utf-8")).hexdigest()[:16]
    return {"version": f"{latest_review_id}:{candidate_digest}"}


# --------------------------------------------------------------------------- #
# Config (exposes the Maps API key to the frontend; never hardcoded)
# --------------------------------------------------------------------------- #
@app.get("/config")
def get_config():
    return {
        "google_maps_api_key": os.environ.get("GOOGLE_MAPS_API_KEY", ""),
    }


# --------------------------------------------------------------------------- #
# Projects
# --------------------------------------------------------------------------- #
@app.post("/projects", response_model=schemas.ProjectOut)
def create_project(
    payload: schemas.ProjectCreate,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_role("sysadmin")),
):
    project = models.Project(
        name=payload.name,
        project_url=payload.project_url,
        notes=payload.notes,
    )
    db.add(project)
    db.commit()
    db.refresh(project)
    return project


@app.get("/projects", response_model=list[schemas.ProjectOut])
def list_projects(
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.get_current_user),
):
    return db.scalars(select(models.Project).order_by(models.Project.created_at.desc())).all()


@app.get("/projects/{project_id}", response_model=schemas.ProjectOut)
def get_project(
    project_id: str,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.get_current_user),
):
    project = db.get(models.Project, project_id)
    if not project:
        raise HTTPException(404, "Project not found")
    return project


def _parse_config(config: Optional[str]) -> schemas.IngestConfig:
    """Accept the ingest config as JSON or YAML in a form field."""
    if not config:
        return schemas.IngestConfig()
    try:
        data = json.loads(config)
    except json.JSONDecodeError:
        try:
            data = yaml.safe_load(config)
        except yaml.YAMLError as exc:  # pragma: no cover
            raise HTTPException(400, f"Could not parse config: {exc}")
    if not isinstance(data, dict):
        raise HTTPException(400, "Config must be a mapping/object")
    return schemas.IngestConfig(**data)


@app.post("/projects/{project_id}/ingest", response_model=schemas.IngestResult)
async def ingest_project(
    project_id: str,
    file: UploadFile = File(...),
    config: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_role("sysadmin")),
):
    project = db.get(models.Project, project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    cfg = _parse_config(config)
    content = await file.read()
    try:
        df = ingestion.read_table(content, file.filename or "upload.csv")
    except Exception as exc:
        raise HTTPException(400, f"Could not read tabular file: {exc}")

    if df.empty:
        raise HTTPException(400, "Source file has no rows")

    # Prefer explicit Latitud/Longitud columns over a single map-reference column.
    lat_col, lng_col = ingestion.detect_latlon_columns(df)
    map_column: Optional[str] = None
    if lat_col and lng_col:
        coord_label = f"{lat_col} + {lng_col}"
        # If the file also has an address column, keep it as map_ref display.
        addr_lowered = {c.lower(): c for c in df.columns}
        map_column = (
            addr_lowered.get("dirección")
            or addr_lowered.get("direccion")
            or addr_lowered.get("address")
            or addr_lowered.get("direccion")
        )
    else:
        try:
            map_column = ingestion.resolve_map_column(df, cfg.map_column)
        except ValueError as exc:
            raise HTTPException(400, str(exc))
        coord_label = map_column
        lat_col = lng_col = None

    records, parsed, failed = ingestion.build_candidates(df, map_column, lat_col, lng_col)

    for rec in records:
        db.add(
            models.LocationCandidate(
                project_id=project_id,
                map_ref=rec["map_ref"],
                lat=rec["lat"],
                lng=rec["lng"],
                display_data=rec["display_data"],
                current_stage=workflow.JEFATURA,
                workflow_group=workflow.PENDING,
            )
        )
    project.source_file = file.filename
    db.commit()

    return schemas.IngestResult(
        project_id=project_id,
        rows_read=len(df),
        candidates_created=len(records),
        map_column=coord_label,
        parsed_coordinates=parsed,
        failed_coordinates=failed,
    )


# --------------------------------------------------------------------------- #
# Review queue + actions (role-scoped)
# --------------------------------------------------------------------------- #
def _review_out(r: models.Review) -> schemas.ReviewOut:
    return schemas.ReviewOut(
        id=r.id,
        candidate_id=r.candidate_id,
        stage=r.stage,
        action=r.action,
        note=r.note,
        created_at=r.created_at,
        reviewer_id=r.reviewer_id,
        reviewer_name=r.reviewer.name if r.reviewer else None,
        reviewer_role=r.reviewer.role if r.reviewer else None,
    )


def _latest_review(
    db: Session,
    candidate_id: int,
    actions: set[str],
    stage: Optional[str] = None,
) -> Optional[models.Review]:
    q = (
        select(models.Review)
        .where(models.Review.candidate_id == candidate_id)
        .where(models.Review.action.in_(actions))
    )
    if stage:
        q = q.where(models.Review.stage == stage)
    q = q.order_by(models.Review.created_at.desc(), models.Review.id.desc()).limit(1)
    return db.scalars(q).first()


def _as_utc_datetime(value: object) -> Optional[datetime]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        dt = value
    else:
        text = str(value).strip()
        if not text:
            return None
        try:
            dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _santiago_datetime(value: object) -> Optional[datetime]:
    dt = _as_utc_datetime(value)
    return dt.astimezone(SANTIAGO_TZ) if dt else None


def _santiago_iso(value: object) -> Optional[str]:
    dt = _santiago_datetime(value)
    return dt.isoformat() if dt else None


def _santiago_display(value: object) -> str:
    dt = _santiago_datetime(value)
    if not dt:
        return "" if value in (None, "") else str(value)
    return dt.strftime("%d-%m-%Y %H:%M")


def _review_date(review: Optional[models.Review]) -> Optional[str]:
    return _santiago_iso(review.created_at) if review else None


def _ensure_review_session_started(request: Request) -> None:
    if not request.session.get("review_session_started_at"):
        request.session["review_session_started_at"] = datetime.now(timezone.utc).isoformat()


DECIDING_ACTIONS_FOR_UI = {"accept", "reject", "study", "project", "like", "dislike", "opening"}


def _candidate_requested_by(candidate: models.LocationCandidate) -> Optional[str]:
    data = candidate.display_data or {}
    email = ""
    for key in ("CorreoSolicitante", "Correo Solicitante", "CORREOSOLICITANTE"):
        if data.get(key):
            email = str(data[key]).strip().lower()
            break
    for category, emails in REQUESTER_CATEGORY_EMAILS.items():
        if email in emails:
            return category
    return None


def _normalized_cut(value: object) -> str:
    text_value = str(value or "").strip().upper()
    if re.fullmatch(r"-?\d+\.0+", text_value):
        return text_value.split(".", 1)[0]
    return text_value


def _attribute_value(attributes: dict, name: str) -> object:
    normalized_name = name.strip().lower()
    for key, value in attributes.items():
        if str(key).strip().lower() == normalized_name:
            return value
    return None


def _candidate_commune_locations(
    db: Session,
    candidate: models.LocationCandidate,
) -> list[dict[str, str]]:
    candidate_cut = _normalized_cut(_attribute_value(candidate.display_data or {}, "CUT"))
    if not candidate_cut:
        return []

    matches: list[dict[str, str]] = []
    for business in db.scalars(select(models.BusinessLocation)).all():
        attributes = business.attributes or {}
        if _attribute_value(attributes, "_source_table") != "LocalesSimi":
            continue
        if _normalized_cut(_attribute_value(attributes, "CUT")) != candidate_cut:
            continue
        matches.append(
            {
                "CveUnidad": str(_attribute_value(attributes, "CveUnidad") or "").strip(),
                "Unidad": str(_attribute_value(attributes, "Unidad") or business.name or "").strip(),
                "Estatus": str(_attribute_value(attributes, "Estatus") or "").strip(),
            }
        )
    return sorted(matches, key=lambda row: (row["CveUnidad"], row["Unidad"]))


def _candidate_out(db: Session, candidate: models.LocationCandidate) -> schemas.CandidateOut:
    group = workflow.candidate_group(db, candidate)
    last_decision = (
        candidate.last_action
        if candidate.last_action in DECIDING_ACTIONS_FOR_UI
        else None
    )
    workflow_dates = {
        "jefatura_like": _santiago_iso(candidate.suggested_at),
        "rejected": _santiago_iso(candidate.rejected_at),
        "observation": _santiago_iso(candidate.rejected_at),
        "study": _santiago_iso(candidate.last_action_at if group == "study" else None),
        "proposed": _santiago_iso(candidate.approved_at),
        "approved": _santiago_iso(candidate.project_at),
        "opening": _santiago_iso(candidate.last_action_at if group == "opening" else None),
    }
    variables = candidate.project_variables
    project_variables = _project_variables_out(candidate.id, variables).model_dump() if variables else None
    approval_conditions = (
        _committee_approval_conditions(db, candidate)
        if group in {"approved", "opening"}
        else None
    )
    current_stage = candidate.current_stage
    if group == "proposed":
        current_stage = workflow.PROPOSED_STAGE
    elif group == "approved":
        current_stage = workflow.APPROVED_STAGE
    elif group == "opening":
        current_stage = workflow.PROJECT_STAGE
    return schemas.CandidateOut(
        id=candidate.id,
        project_id=candidate.project_id,
        map_ref=candidate.map_ref,
        lat=candidate.lat,
        lng=candidate.lng,
        display_data=candidate.display_data or {},
        requested_by=_candidate_requested_by(candidate),
        current_stage=current_stage,
        status=candidate.status,
        workflow_group=group,
        priority=candidate.priority,
        last_decision=last_decision,
        last_reject_note=candidate.last_reject_note,
        workflow_dates=workflow_dates,
        project_variables=project_variables,
        approved_division=(
            _committee_selected_division(db, candidate)
            if group in {"approved", "opening"}
            else _candidate_source_division(candidate)
        ),
        approval_conditions=approval_conditions,
    )


def _stats_payload(db: Session, project_id: Optional[str] = None) -> dict:
    q = select(models.LocationCandidate)
    if project_id:
        q = q.where(models.LocationCandidate.project_id == project_id)

    queues = {stage: 0 for stage in workflow.STAGES}
    statuses = {
        "pending": 0,
        "returned": 0,
        "rejected": 0,
        "observation": 0,
        "study": 0,
        "proposed": 0,
        "approved": 0,
        "por_abrir": 0,
    }
    total = 0
    for candidate in db.scalars(q).all():
        total += 1
        group = workflow.candidate_group(db, candidate)
        if group == "pending":
            statuses["pending"] += 1
            queues["jefatura"] += 1
            queues["jefecomercial"] += 1
            queues["coordinador"] += 1
            queues["arriendo"] += 1
            queues["gerente"] += 1
        elif group == "proposed":
            statuses["proposed"] += 1
            queues["comite"] += 1
            queues["gerentegeneral"] += 1
        elif group == "rejected":
            statuses["rejected"] += 1
        elif group == "observation":
            statuses["observation"] += 1
        elif group == "study":
            statuses["study"] += 1
        elif group == "approved":
            statuses["approved"] += 1
            queues["coordinador"] += 1
        elif group == "opening":
            statuses["por_abrir"] += 1
    return {"total": total, "queues": queues, "statuses": statuses}


def _display_text(data: dict, keys: list[str]) -> str:
    value = _display_value(data, keys)
    return str(value or "").strip()


def _candidate_projection_email(candidate: models.LocationCandidate) -> str:
    return _display_text(
        candidate.display_data or {},
        ["CorreoSolicitante", "Correo Solicitante", "CORREOSOLICITANTE"],
    ).lower()


def _candidate_source_division(candidate: models.LocationCandidate) -> str:
    return _display_text(candidate.display_data or {}, ["DIVISION", "Division", "División"]).upper()


def _email_list(value: Optional[str]) -> set[str]:
    if not value:
        return set()
    return {
        part.strip().lower()
        for chunk in value.replace(";", "\n").replace(",", "\n").splitlines()
        for part in [chunk]
        if part.strip()
    }


def _division_from_note(note: Optional[str]) -> Optional[str]:
    if not note:
        return None
    # Anchor to the "División:" label so free-text approval conditions in the
    # same note can't be mistaken for the division (a bare substring match would).
    matches = re.findall(r"DIVISI[OÓ]N\s*:\s*(SUCURSAL|FRANQUICIA)", note.upper())
    return matches[-1] if matches else None


def _require_current_approval_division(
    db: Session,
    candidate: models.LocationCandidate,
    action: str,
    note: Optional[str],
) -> None:
    if workflow.candidate_group(db, candidate) != "proposed" or action not in {"accept", "project"}:
        return
    if not _division_from_note(note):
        raise HTTPException(
            400,
            "Seleccione Sucursal o Franquicia para esta aprobación.",
        )


def _conditions_from_note(note: Optional[str]) -> Optional[str]:
    """Free-text approval conditions the committee left, if any."""
    if not note:
        return None
    match = re.search(
        r"condiciones de aprobaci[oó]n\s*:\s*(.+)",
        note,
        re.IGNORECASE | re.DOTALL,
    )
    if not match:
        return None
    return match.group(1).strip() or None


def _committee_approval_conditions(
    db: Session, candidate: models.LocationCandidate
) -> Optional[str]:
    """Conditions from the committee's final decision (the latest 'project' review)."""
    review = db.scalars(
        select(models.Review)
        .where(models.Review.candidate_id == candidate.id)
        .where(models.Review.action == "project")
        .order_by(models.Review.created_at.desc(), models.Review.id.desc())
        .limit(1)
    ).first()
    return _conditions_from_note(review.note if review else None)


def _committee_selected_division(db: Session, candidate: models.LocationCandidate) -> str:
    # The division chosen at the committee's final decision wins. That decision
    # is always recorded as a "project" action -- by the committee, the general
    # manager, or a sysadmin override -- regardless of the reviewer's stage. Fall
    # back to the legacy approver-stage note so locations approved before this
    # change keep resolving, then to the source division.
    for conditions in (
        (models.Review.action == "project",),
        (
            models.Review.stage.in_(tuple(workflow.APPROVER_ROLES)),
            models.Review.action == "accept",
        ),
    ):
        review = db.scalars(
            select(models.Review)
            .where(models.Review.candidate_id == candidate.id)
            .where(*conditions)
            .order_by(models.Review.created_at.desc(), models.Review.id.desc())
            .limit(1)
        ).first()
        division = _division_from_note(review.note if review else None)
        if division:
            return division
    return _candidate_source_division(candidate)


def _normal_commercial_division(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    normalized = value.strip().upper()
    return normalized if normalized in COMMERCIAL_DIVISIONS else None


def _normal_jefatura_group(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    normalized = value.strip().upper()
    return normalized if normalized in JEFATURA_GROUPS else None


def _candidate_visible_to_user(
    db: Session,
    candidate: models.LocationCandidate,
    user: models.User,
    commercial_division: Optional[str] = None,
) -> bool:
    if user.role in {
        workflow.SYSADMIN,
        workflow.COMITE,
        workflow.GERENTE_GENERAL,
        workflow.GERENTE,
        workflow.ARRIENDO,
    }:
        return True
    if user.role == workflow.JEFATURA:
        if user.email.lower() == JEFATURA_ADMIN_EMAIL:
            return True
        selected = _normal_jefatura_group(user.commercial_division)
        if selected == "APERTURA":
            return True
        if selected in COMMERCIAL_DIVISIONS:
            return _candidate_source_division(candidate) == selected
        return _candidate_projection_email(candidate) == user.email.lower()
    if user.role == workflow.JEFE_COMERCIAL:
        selected = _normal_commercial_division(user.commercial_division)
        if not selected:
            return False
        candidate_email = _candidate_projection_email(candidate)
        is_own_candidate = candidate_email == user.email.lower()
        supervisor_emails = _email_list(user.supervisor_emails)
        if not is_own_candidate and candidate_email not in supervisor_emails:
            return False
        if workflow.candidate_group(db, candidate) in {"proposed", "approved", "opening"}:
            return _committee_selected_division(db, candidate) == selected
        return _candidate_source_division(candidate) == selected
    if user.role == workflow.COORDINADOR:
        selected = _normal_commercial_division(user.commercial_division)
        if not selected:
            return False
        # Approved/project locations follow the division chosen by the committee,
        # so a location proposed for one division but approved for another is
        # visible to the coordinator of the approved division.
        if workflow.candidate_group(db, candidate) in {"approved", "opening"}:
            return _committee_selected_division(db, candidate) == selected
        return _candidate_source_division(candidate) == selected
    return False


def _visible_candidates(
    db: Session,
    user: models.User,
    project_id: Optional[str] = None,
    commercial_division: Optional[str] = None,
) -> list[models.LocationCandidate]:
    q = select(models.LocationCandidate).order_by(models.LocationCandidate.id)
    if project_id:
        q = q.where(models.LocationCandidate.project_id == project_id)
    return [
        candidate
        for candidate in db.scalars(q).all()
        if _candidate_visible_to_user(db, candidate, user, commercial_division)
    ]


def _require_candidate_visible(
    db: Session,
    candidate: models.LocationCandidate,
    user: models.User,
    commercial_division: Optional[str] = None,
) -> None:
    if not _candidate_visible_to_user(db, candidate, user, commercial_division):
        raise HTTPException(403, "Candidate is outside this user's scope.")


def _queue_visible_candidates(
    db: Session,
    user: models.User,
    project_id: Optional[str] = None,
    commercial_division: Optional[str] = None,
) -> list[models.LocationCandidate]:
    visible = _visible_candidates(db, user, project_id, commercial_division)
    if user.role not in workflow.JEFATURA_LIKE_ROLES:
        return visible
    reviewed_ids = set(
        db.scalars(
            select(models.Review.candidate_id)
            .where(models.Review.reviewer_id == user.id)
            .where(models.Review.action.in_({"like", "dislike"}))
        ).all()
    )
    if not reviewed_ids:
        return visible
    return [candidate for candidate in visible if candidate.id not in reviewed_ids]


def _action_out(
    db: Session,
    candidate: models.LocationCandidate,
    user: models.User,
    project_id: Optional[str] = None,
    sort_by: str = "score",
    sort_dir: str = "desc",
    commercial_division: Optional[str] = None,
) -> schemas.CandidateActionOut:
    visible = _queue_visible_candidates(db, user, project_id, commercial_division)
    next_items = workflow.candidates_for_role(
        db, user.role, project_id, sort_by, sort_dir, candidates=visible
    )
    next_candidate = next((c for c in next_items if c.id != candidate.id), None)
    return schemas.CandidateActionOut(
        candidate=_candidate_out(db, candidate),
        next_candidate=_candidate_out(db, next_candidate) if next_candidate else None,
        remaining=len(next_items),
        stats=_stats_payload(db, project_id),
    )


@app.get("/queue", response_model=schemas.QueueOut)
def get_queue(
    project_id: Optional[str] = None,
    sort_by: str = "score",
    sort_dir: str = "desc",
    division: Optional[str] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    """Next candidate in the current user's review layer (sysadmin has none)."""
    stage = workflow.role_stage(user.role)
    if stage is None:
        return schemas.QueueOut(candidate=None, remaining=0, stage=None)

    visible = _queue_visible_candidates(db, user, project_id, division)
    candidates = workflow.candidates_for_role(
        db, user.role, project_id, sort_by, sort_dir, candidates=visible
    )
    remaining = len(candidates)
    candidate = candidates[0] if candidates else None
    return schemas.QueueOut(
        candidate=_candidate_out(db, candidate) if candidate else None,
        remaining=remaining,
        stage=stage,
    )


@app.get("/candidates", response_model=list[schemas.CandidateOut])
def list_candidates(
    project_id: Optional[str] = None,
    division: Optional[str] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    return [_candidate_out(db, c) for c in _visible_candidates(db, user, project_id, division)]


@app.get("/candidates/by-projection/{projection_id}", response_model=schemas.CandidateOut)
def get_pending_candidate_by_projection(
    projection_id: str,
    division: Optional[str] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    candidate = _candidate_by_projection_id(db, projection_id)
    if not candidate:
        raise HTTPException(404, "Projection ID not found.")
    _require_candidate_visible(db, candidate, user, division)
    if workflow.candidate_group(db, candidate) != "pending":
        raise HTTPException(409, "Projection ID is not pending.")
    return _candidate_out(db, candidate)


@app.get("/candidates/by-projection/{projection_id}/audit")
def candidate_audit_by_projection(
    projection_id: str,
    division: Optional[str] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    candidate = _candidate_by_projection_id(db, projection_id)
    if not candidate:
        raise HTTPException(404, "Projection ID not found.")
    _require_candidate_visible(db, candidate, user, division)
    return _candidate_audit_payload(db, candidate)


EXPORT_GROUPS = {
    "pending": "Pendientes",
    "observation": "Observación",
    "rejected": "Rechazados",
    "study": "En Estudio",
    "proposed": "Propuestos",
    "approved": "Aprobados",
    "opening": "Proyectos",
}

EXPORT_FILE_SLUGS = {
    "pending": "pendientes",
    "observation": "observacion",
    "rejected": "rechazados",
    "study": "en_estudio",
    "proposed": "propuestos",
    "approved": "aprobados",
    "opening": "proyectos",
}

PROJECT_VARIABLE_EXPORT_COLUMNS = [
    ("cve_unidad", "CveUnidad"),
    ("unidad", "Unidad"),
    ("comuna", "Comuna"),
    ("provincia", "Provincia"),
    ("region", "Region"),
    ("mt2", "MT2"),
    ("valor_arriendo", "Valor de Arriendo"),
    ("gastos_comunes", "Gastos Comunes"),
    ("clausula_salida", "Clausula de salida"),
    ("meses_gracia", "Meses de gracia"),
    ("plazo_arriendo", "Plazo de arriendo"),
    ("garantia", "Garantia"),
    ("tipo_proyecto", "Tipo de Proyecto"),
    ("fecha_apertura_aproximada", "Fecha aproximada de apertura"),
    ("contacto_nombre", "Nombre contacto"),
    ("contacto_telefono", "Telefono contacto"),
    ("contacto_email", "Email contacto"),
    ("flujo_franquicia", "Flujo Franquicia"),
    ("franquiciado_nombre", "Nombre franquiciado"),
    ("franquiciado_telefono", "Telefono franquiciado"),
    ("franquiciado_email", "Email franquiciado"),
    ("fecha_entrega_local", "Fecha entrega local"),
]


def _candidate_export_group(db: Session, candidate: models.LocationCandidate) -> str:
    return workflow.candidate_group(db, candidate)


def _project_variables_out(
    candidate_id: int,
    variables: Optional[models.CandidateProjectVariables],
) -> schemas.CandidateProjectVariablesOut:
    if variables is None:
        return schemas.CandidateProjectVariablesOut(candidate_id=candidate_id)

    def text_or_none(value: object) -> Optional[str]:
        if value in (None, ""):
            return None
        return str(value)

    return schemas.CandidateProjectVariablesOut(
        candidate_id=candidate_id,
        cve_unidad=variables.cve_unidad,
        unidad=variables.unidad,
        comuna=variables.comuna,
        provincia=variables.provincia,
        region=variables.region,
        mt2=variables.mt2,
        valor_arriendo=text_or_none(variables.valor_arriendo),
        gastos_comunes=text_or_none(variables.gastos_comunes),
        clausula_salida=variables.clausula_salida,
        meses_gracia=text_or_none(variables.meses_gracia),
        plazo_arriendo=variables.plazo_arriendo,
        garantia=variables.garantia,
        tipo_proyecto=variables.tipo_proyecto,
        fecha_apertura_aproximada=variables.fecha_apertura_aproximada,
        contacto_nombre=variables.contacto_nombre,
        contacto_telefono=variables.contacto_telefono,
        contacto_email=variables.contacto_email,
        flujo_franquicia=variables.flujo_franquicia,
        franquiciado_nombre=variables.franquiciado_nombre,
        franquiciado_telefono=variables.franquiciado_telefono,
        franquiciado_email=variables.franquiciado_email,
        fecha_entrega_local=variables.fecha_entrega_local,
        updated_at=variables.updated_at,
        updated_by_id=variables.updated_by_id,
    )


def _clean_project_variables_payload(
    payload: schemas.CandidateProjectVariablesIn,
) -> dict:
    values = payload.model_dump()
    for key, value in list(values.items()):
        if isinstance(value, str):
            value = value.strip()
            values[key] = value.upper() if value else None
    for key in ("cve_unidad", "unidad"):
        if values.get(key):
            values[key] = str(values[key]).upper()
        else:
            raise HTTPException(400, "CveUnidad y Unidad son obligatorios.")
    return values


def _project_email_value(value: object, fallback: str = "") -> str:
    if value in (None, ""):
        return fallback
    return str(value)


def _project_email_date(value: object, fallback: str = "") -> str:
    if value in (None, ""):
        return fallback
    if isinstance(value, (date, datetime)):
        return value.strftime("%d-%m-%Y")
    raw = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, pattern).strftime("%d-%m-%Y")
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).strftime("%d-%m-%Y")
    except ValueError:
        return raw


def _project_email_context(
    candidate: models.LocationCandidate,
    values: dict,
) -> dict[str, str]:
    data = candidate.display_data or {}
    projection_id = _display_value(data, ["ID Proyección", "ID Proyeccion", "ID"]) or candidate.id
    address = _display_value(data, ["DIRECCIÓN", "DIRECCION", "Direccion", "DIRECCIÃ“N"]) or candidate.map_ref or ""
    comuna = values.get("comuna") or _display_value(data, ["Comuna", "COMUNA"])
    provincia = values.get("provincia") or _display_value(data, ["Provincia", "PROVINCIA"])
    region = values.get("region") or _display_value(data, ["Region", "REGION"])
    mt2 = _project_email_value(values.get("mt2") or data.get("MT2"))
    delivery_date = _project_email_date(
        values.get("fecha_entrega_local") or values.get("fecha_apertura_aproximada"),
        "A PARTIR DE LA FIRMA DE CONTRATO",
    )
    return {
        "projection_id": _project_email_value(projection_id),
        "delivery_date": delivery_date,
        "cve_unidad": _project_email_value(values.get("cve_unidad")),
        "unidad": _project_email_value(values.get("unidad")),
        "address": _project_email_value(address),
        "comuna": _project_email_value(comuna),
        "provincia": _project_email_value(provincia),
        "region": _project_email_value(region),
        "mt2": f"{mt2} MT2" if mt2 else "",
        "valor_arriendo": _project_email_value(values.get("valor_arriendo")),
        "gastos_comunes": _project_email_value(values.get("gastos_comunes")),
        "clausula_salida": _project_email_value(values.get("clausula_salida")),
        "meses_gracia": _project_email_value(values.get("meses_gracia")),
        "plazo_arriendo": _project_email_value(values.get("plazo_arriendo")),
        "garantia": _project_email_value(values.get("garantia")),
        "contact_name": _project_email_value(values.get("contacto_nombre")),
        "contact_phone": _project_email_value(values.get("contacto_telefono")),
        "contact_email": _project_email_value(values.get("contacto_email")),
        "franchisee_name": _project_email_value(values.get("franquiciado_nombre")),
        "franchisee_phone": _project_email_value(values.get("franquiciado_telefono")),
        "franchisee_email": _project_email_value(values.get("franquiciado_email")),
    }


def _project_email_body(
    candidate: models.LocationCandidate,
    values: dict,
    reduced: bool = False,
    include_franchisee: bool = False,
    signature_name: str = "",
    signature_title: str = "",
) -> str:
    ctx = _project_email_context(candidate, values)
    lines = [
        "Estimados, buen dia",
        "",
        "Dejo el contacto nuevo proyecto para gestion de cada area",
        f"El ID de proyeccion es #{ctx['projection_id']}",
        "Area de Aperturas y Remodelacion, su apoyo con factibilidad y desarrollo de proyectos.",
        "",
        f"ENTREGA DE PROYECTO FECHA APROXIMADA: {ctx['delivery_date']}",
        "",
        "LOCAL PARA PROYECTO NUEVO",
    ]

    def add_value(label: str, key: str) -> None:
        if ctx[key]:
            lines.append(f"{label}: {ctx[key]}")

    for label, key in (
        ("UNIDAD", "cve_unidad"),
        ("NOMBRE", "unidad"),
        ("DIRECCION", "address"),
        ("COMUNA", "comuna"),
        ("PROVINCIA", "provincia"),
        ("REGION", "region"),
        ("MT2 LOCAL", "mt2"),
    ):
        add_value(label, key)
    if not reduced:
        for label, key in (
            ("VALOR", "valor_arriendo"),
            ("GGCC", "gastos_comunes"),
            ("CLAUSULA SALIDA MES A FAVOR DE SIMI", "clausula_salida"),
            ("MESES DE GRACIA", "meses_gracia"),
            ("PLAZO DE ARRIENDO", "plazo_arriendo"),
            ("GARANTIA", "garantia"),
        ):
            add_value(label, key)
    if any(ctx[key] for key in ("contact_name", "contact_phone", "contact_email")):
        lines.extend(["", "CONTACTO"])
        for label, key in (("NOMBRE", "contact_name"), ("TELEFONO", "contact_phone"), ("EMAIL", "contact_email")):
            add_value(label, key)
    if (not reduced or include_franchisee) and any(
        ctx[key] for key in ("franchisee_name", "franchisee_phone", "franchisee_email")
    ):
        lines.extend(["", "FRANQUICIADO"])
        for label, key in (("NOMBRE", "franchisee_name"), ("TELEFONO", "franchisee_phone"), ("EMAIL", "franchisee_email")):
            add_value(label, key)
    lines.extend(["", "Saludos,"])
    if signature_name:
        lines.append(signature_name)
    if signature_title:
        lines.append(signature_title)
    return "\n".join(lines)


def _project_email_html_table(
    ctx: dict[str, str],
    reduced: bool = False,
    include_franchisee: bool = False,
) -> str:
    def e(key: str) -> str:
        return html_escape(ctx.get(key, ""))

    def row(label: str, value: str, underline: bool = False) -> str:
        if not value:
            return ""
        value_style = "padding:6px; border:1px solid black;"
        if underline:
            value_style += " text-decoration: underline;"
        return (
            "<tr>"
            '<td style="background:#D9E1F2; color:black; padding:6px; border:1px solid black; white-space: nowrap;">'
            f"<b>{html_escape(label)}</b></td>"
            f'<td style="{value_style}">{value}</td>'
            "</tr>"
        )

    email = e("contact_email")
    email_html = (
        f'<a href="mailto:{email}" style="color:#0070C0;">{email}</a>'
        if email
        else ""
    )
    table = (
        '<table style="border-collapse: collapse; border-spacing:0; width: auto; '
        'font-family: Arial, sans-serif; font-size:12px; border:1px solid black; '
        'table-layout: auto; mso-table-lspace:0pt; mso-table-rspace:0pt;">'
        "<thead><tr>"
        '<th colspan="2" style="background-color:#D9E1F2; color:black; padding:10px; border:1px solid black; text-align:center;">'
        "<b>LOCAL PARA PROYECTO NUEVO</b></th>"
        "</tr></thead><tbody>"
        + row("UNIDAD", e("cve_unidad"))
        + row("NOMBRE", e("unidad"))
        + row("DIRECCIÓN", e("address"))
        + row("COMUNA", e("comuna"))
        + row("PROVINCIA", e("provincia"), underline=True)
        + row("REGIÓN", e("region"), underline=True)
        + row("MTS2 LOCAL", e("mt2"))
    )
    if not reduced:
        table += (
            row("VALOR", e("valor_arriendo"))
            + row("GGCC", e("gastos_comunes"))
            + row("CLAUSULA SALIDA MES A FAVOR DE SIMI", e("clausula_salida"))
            + row("MESES DE GRACIA", e("meses_gracia"))
            + row("PLAZO DE ARRIENDO", e("plazo_arriendo"))
            + row("GARANTIA", e("garantia"))
        )
    if any(ctx[key] for key in ("contact_name", "contact_phone", "contact_email")):
        table += (
            '<tr><td colspan="2" style="background:#D9E1F2; color:black; padding:10px; border:1px solid black; text-align:center;"><b>CONTACTO</b></td></tr>'
            + row("NOMBRE", e("contact_name"))
            + row("TELÉFONO", e("contact_phone"))
            + row("EMAIL", email_html)
        )
    if (not reduced or include_franchisee) and any(
        ctx[key] for key in ("franchisee_name", "franchisee_phone", "franchisee_email")
    ):
        franchisee_email = e("franchisee_email")
        franchisee_email_html = (
            f'<a href="mailto:{franchisee_email}" style="color:#0070C0;">{franchisee_email}</a>'
            if franchisee_email
            else ""
        )
        table += (
            '<tr><td colspan="2" style="background:#D9E1F2; color:black; padding:10px; border:1px solid black; text-align:center;"><b>FRANQUICIADO</b></td></tr>'
            + row("NOMBRE", e("franchisee_name"))
            + row("TELÉFONO", e("franchisee_phone"))
            + row("EMAIL", franchisee_email_html)
        )
    return table + "</tbody></table>"


def _project_email_html_body(
    candidate: models.LocationCandidate,
    values: dict,
    reduced: bool = False,
    include_franchisee: bool = False,
    signature_name: str = "",
    signature_title: str = "",
) -> str:
    ctx = _project_email_context(candidate, values)
    signature_html = "".join(
        f"<br>{html_escape(line)}"
        for line in (signature_name, signature_title)
        if line
    )
    return f"""\
<html>
  <body style="font-family: Arial, sans-serif; font-size:14px; color:#222;">
    <p>Estimados, buen día</p>
    <p>Dejo el contacto nuevo proyecto para gestión de cada área</p>
    <p>El ID de proyección es #{html_escape(ctx['projection_id'])}<br>
    Área de Aperturas y Remodelación, su apoyo con factibilidad y desarrollo de proyectos.</p>
    <p>ENTREGA DE PROYECTO FECHA APROXIMADA: {html_escape(ctx['delivery_date'])}</p>
    {_project_email_html_table(ctx, reduced=reduced, include_franchisee=include_franchisee)}
    <p>Saludos,{signature_html}</p>
  </body>
</html>
"""


def _project_email_subject(values: dict) -> str:
    return f"NUEVO LOCAL APROBADO {values.get('cve_unidad') or ''} {values.get('unidad') or ''}".strip()


def _project_email_plans(
    db: Session,
    candidate: models.LocationCandidate,
    values: dict,
) -> list[dict]:
    division = _committee_selected_division(db, candidate).upper()
    subject = _project_email_subject(values)
    if division == "SUCURSAL":
        signature_name = "Jennifer Villavicencio"
        signature_title = "Coordinadora de Proyecto"
    elif division == "FRANQUICIA":
        signature_name = "Leonel Albornoz"
        signature_title = "Coordinador de proyectos franquicias"
    else:
        signature_name = ""
        signature_title = ""

    def plan(
        plan_id: str,
        area: str,
        from_email: str,
        recipients: list[str],
        cc: list[str],
        reduced: bool,
        include_franchisee: bool = False,
    ) -> dict:
        return {
            "plan_id": plan_id,
            "area": area,
            "from_email": from_email,
            "recipients": recipients,
            "cc": cc,
            "subject": subject,
            "html_body": _project_email_html_body(
                candidate,
                values,
                reduced=reduced,
                include_franchisee=include_franchisee,
                signature_name=signature_name,
                signature_title=signature_title,
            ),
            "reduced": reduced,
            "include_franchisee": include_franchisee,
            "signature_name": signature_name,
            "signature_title": signature_title,
        }

    if division == "SUCURSAL":
        return [
            plan(
                "sucursal_legal",
                "Arriendo y Legal",
                SUCURSAL_ORIGIN_EMAIL,
                SUCURSAL_LEGAL_TO,
                SUCURSAL_LEGAL_CC,
                False,
            ),
            plan(
                "sucursal_reducido",
                "Arquitectura y áreas internas",
                SUCURSAL_ORIGIN_EMAIL,
                SUCURSAL_REDUCED_TO,
                SUCURSAL_REDUCED_CC,
                True,
            ),
        ]

    if division != "FRANQUICIA":
        raise HTTPException(400, "El local debe estar aprobado como Sucursal o Franquicia.")
    missing_contact = [
        label
        for key, label in (
            ("contacto_nombre", "Nombre del contacto"),
            ("contacto_telefono", "Teléfono del contacto"),
            ("contacto_email", "Email del contacto"),
        )
        if not values.get(key)
    ]
    if missing_contact:
        raise HTTPException(400, f"Complete: {', '.join(missing_contact)}.")

    franchise_flow = str(values.get("flujo_franquicia") or "").upper()
    if franchise_flow == "FRANQUICIADO DIRECTO":
        applicant_email = _candidate_projection_email(candidate)
        return [
            plan(
                "franquicia_directa",
                "Franquiciado Directo",
                FRANCHISE_ORIGIN_EMAIL,
                ["mbustos@farmaciasdoctorsimi.cl", "rmalave@farmaciasdoctorsimi.cl"],
                [applicant_email] if applicant_email else [],
                False,
            )
        ]
    if franchise_flow == "SUBARRIENDO":
        return [
            plan(
                "subarriendo_legal",
                "Legal",
                FRANCHISE_ORIGIN_EMAIL,
                ["cfolsch@farmaciasdoctorsimi.cl", "curibe@farmaciasdoctorsimi.cl"],
                [],
                False,
            ),
            plan(
                "subarriendo_arquitectura",
                "Arquitectura",
                FRANCHISE_ORIGIN_EMAIL,
                ["ptarsetti@farmaciasdoctorsimi.cl"],
                [],
                True,
                include_franchisee=True,
            ),
        ]
    raise HTTPException(400, "Seleccione Subarriendo o Franquiciado Directo.")


def _send_project_email_plan(
    candidate: models.LocationCandidate,
    values: dict,
    plan: dict,
    recipients: list[str],
    cc: list[str],
) -> schemas.CandidateProjectSentEmailOut:
    msg = EmailMessage()
    msg["From"] = plan["from_email"]
    msg["To"] = ", ".join(recipients)
    if cc:
        msg["Cc"] = ", ".join(cc)
    msg["Subject"] = plan["subject"]
    msg.set_content(
        _project_email_body(
            candidate,
            values,
            reduced=plan["reduced"],
            include_franchisee=plan["include_franchisee"],
            signature_name=plan["signature_name"],
            signature_title=plan["signature_title"],
        )
    )
    msg.add_alternative(plan["html_body"], subtype="html")

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=20) as smtp:
            smtp.send_message(
                msg,
                from_addr=plan["from_email"],
                to_addrs=recipients + cc,
            )
    except OSError as exc:
        raise HTTPException(502, f"No se pudo enviar el correo por SMTP: {exc}") from exc
    return schemas.CandidateProjectSentEmailOut(
        plan_id=plan["plan_id"],
        area=plan["area"],
        from_email=plan["from_email"],
        recipients=recipients,
        cc=cc,
        subject=plan["subject"],
    )


def _normalize_project_email_addresses(addresses: list[str], label: str) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for raw_value in addresses:
        for address in re.split(r"[;,\s]+", str(raw_value or "").strip()):
            if not address:
                continue
            key = address.lower()
            if len(address) > 254 or not re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", address):
                raise HTTPException(400, f"Correo no válido en {label}: {address}")
            if key not in seen:
                seen.add(key)
                normalized.append(address)
    if len(normalized) > 50:
        raise HTTPException(400, f"Se permiten hasta 50 correos en {label}.")
    return normalized


def _ensure_project_variables_allowed(
    db: Session,
    candidate: models.LocationCandidate,
    user: models.User,
) -> None:
    if user.role not in {workflow.COORDINADOR, workflow.ARRIENDO, workflow.SYSADMIN}:
        raise HTTPException(403, "Only Coordinador, Arriendos y Patentes, or Sysadmin can edit project variables.")
    if workflow.candidate_group(db, candidate) not in {"approved", "opening"}:
        raise HTTPException(409, "Project variables are only available for Aprobados or Proyectos.")


def _ensure_franchise_activation_variables(
    db: Session,
    candidate: models.LocationCandidate,
) -> None:
    if _committee_selected_division(db, candidate).upper() != "FRANQUICIA":
        return
    variables = candidate.project_variables
    missing = [
        label
        for attr, label in (
            ("contacto_nombre", "Nombre del contacto"),
            ("contacto_telefono", "Teléfono del contacto"),
            ("contacto_email", "Email del contacto"),
        )
        if not (getattr(variables, attr, None) if variables else None)
    ]
    if missing:
        raise HTTPException(409, "Complete Variables antes de dar de alta: " + ", ".join(missing))


def _display_value(display_data: dict, keys: list[str]) -> object:
    for key in keys:
        value = display_data.get(key)
        if value not in (None, ""):
            return value
    return ""


def _candidate_source_date(display_data: dict) -> Optional[datetime]:
    value = _display_value(display_data, ["FECHA", "Fecha", "fecha"])
    return _as_utc_datetime(value)


def _is_date_key(key: str) -> bool:
    return "fecha" in key.lower()


def _candidate_view_date(db: Session, candidate: models.LocationCandidate, group: str) -> object:
    if group == "pending":
        return _santiago_display(_display_value(candidate.display_data or {}, ["FECHA", "Fecha", "fecha"]))
    if group == "proposed":
        review = _latest_review(db, candidate.id, {"accept"})
        return _santiago_display(review.created_at) if review else ""
    if group in {"rejected", "observation"}:
        review = _latest_review(db, candidate.id, {"reject"})
        return _santiago_display(review.created_at if review else candidate.rejected_at)
    if group == "study":
        review = _latest_review(db, candidate.id, {"study"})
        return _santiago_display(review.created_at) if review else ""
    if group == "approved":
        review = _latest_review(db, candidate.id, {"project"})
        return _santiago_display(review.created_at) if review else ""
    if group == "opening":
        review = _latest_review(db, candidate.id, {"opening"}, workflow.COORDINADOR)
        return _santiago_display(review.created_at) if review else ""
    return ""


def _export_rows(
    db: Session,
    candidates: list[models.LocationCandidate],
) -> tuple[list[str], list[list[object]]]:
    display_keys: list[str] = []
    for candidate in candidates:
        for key in candidate.display_data or {}:
            if key not in display_keys:
                display_keys.append(key)

    review_cols: list[str] = []
    for stage in workflow.STAGES:
        review_cols += [f"{stage}_accion", f"{stage}_comentario"]

    project_variable_headers = [label for _, label in PROJECT_VARIABLE_EXPORT_COLUMNS]

    header = [
        "candidate_id",
        "grupo",
        "fecha_vista",
        "estado",
        "etapa_actual",
        "prioridad",
        "latitud",
        "longitud",
        "map_ref",
    ] + review_cols + project_variable_headers + display_keys

    rows: list[list[object]] = []
    for candidate in candidates:
        group = _candidate_export_group(db, candidate)
        row: list[object] = [
            candidate.id,
            EXPORT_GROUPS.get(group, group),
            _candidate_view_date(db, candidate, group),
            candidate.status,
            candidate.current_stage,
            "Si" if candidate.priority else "",
            candidate.lat if candidate.lat is not None else "",
            candidate.lng if candidate.lng is not None else "",
            candidate.map_ref or "",
        ]
        for stage in workflow.STAGES:
            decision = workflow.current_decision(db, candidate.id, stage)
            row += [decision.action if decision else "", (decision.note if decision else "") or ""]
        variables = candidate.project_variables
        for attr, _label in PROJECT_VARIABLE_EXPORT_COLUMNS:
            value = getattr(variables, attr, "") if variables else ""
            row.append("" if value is None else value)
        for key in display_keys:
            value = (candidate.display_data or {}).get(key, "")
            if _is_date_key(key):
                value = _santiago_display(value)
            row.append("" if value is None else value)
        rows.append(row)
    return header, rows


def _add_export_sheet(
    wb: Workbook,
    title: str,
    db: Session,
    candidates: list[models.LocationCandidate],
):
    ws = wb.create_sheet(title=title[:31])
    header, rows = _export_rows(db, candidates)
    ws.append(header)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 12
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 62)


def _add_review_session_sheet(
    wb: Workbook,
    db: Session,
    reviews: list[models.Review],
):
    candidates = [r.candidate for r in reviews if r.candidate is not None]
    display_keys: list[str] = []
    for candidate in candidates:
        for key in candidate.display_data or {}:
            if key not in display_keys:
                display_keys.append(key)

    header = [
        "fecha_accion",
        "accion",
        "comentario",
        "usuario",
        "candidate_id",
        "grupo_actual",
        "estado_actual",
        "etapa_actual",
        "latitud",
        "longitud",
        "map_ref",
    ] + display_keys

    ws = wb.create_sheet(title="Sesion Comite")
    ws.append(header)
    for review in reviews:
        candidate = review.candidate
        if candidate is None:
            continue
        row: list[object] = [
            _santiago_display(review.created_at),
            "Aprobado" if review.action == "accept" else "Rechazado",
            review.note or "",
            review.reviewer.name if review.reviewer else "",
            candidate.id,
            EXPORT_GROUPS.get(_candidate_export_group(db, candidate), _candidate_export_group(db, candidate)),
            candidate.status,
            candidate.current_stage,
            candidate.lat if candidate.lat is not None else "",
            candidate.lng if candidate.lng is not None else "",
            candidate.map_ref or "",
        ]
        for key in display_keys:
            value = (candidate.display_data or {}).get(key, "")
            if _is_date_key(key):
                value = _santiago_display(value)
            row.append("" if value is None else value)
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 12
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 62)


def _export_timestamp() -> str:
    return datetime.now(SANTIAGO_TZ).strftime("%Y%m%d")


@app.get("/candidates/export.xlsx")
def export_candidates_xlsx(
    group: Optional[str] = None,
    all_groups: bool = False,
    project_id: Optional[str] = None,
    division: Optional[str] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    if group and group not in EXPORT_GROUPS:
        raise HTTPException(400, "Invalid export group.")

    candidates = _visible_candidates(db, user, project_id, division)

    wb = Workbook()
    wb.remove(wb.active)

    if all_groups:
        for export_group, label in EXPORT_GROUPS.items():
            sheet_candidates = [c for c in candidates if _candidate_export_group(db, c) == export_group]
            _add_export_sheet(wb, label, db, sheet_candidates)
        filename = f"locales_todas_las_vistas_{_export_timestamp()}.xlsx"
    else:
        export_group = group or "pending"
        label = EXPORT_GROUPS[export_group]
        sheet_candidates = [c for c in candidates if _candidate_export_group(db, c) == export_group]
        _add_export_sheet(wb, label, db, sheet_candidates)
        filename = f"locales_{EXPORT_FILE_SLUGS[export_group]}_{_export_timestamp()}.xlsx"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/candidates/export-session.xlsx")
def export_committee_session_xlsx(
    request: Request,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    if user.role not in workflow.COMITE_LIKE_ROLES:
        raise HTTPException(403, "Only Comité or Gerente General can export a review session.")

    started_at = _as_utc_datetime(request.session.get("review_session_started_at"))
    if started_at is None:
        started_at = datetime.now(timezone.utc)
        request.session["review_session_started_at"] = started_at.isoformat()

    reviews = db.scalars(
        select(models.Review)
        .where(models.Review.reviewer_id == user.id)
        .where(models.Review.stage == user.role)
        .where(models.Review.action.in_({"project", "reject"}))
        .where(models.Review.created_at >= started_at)
        .order_by(models.Review.created_at, models.Review.id)
    ).all()

    wb = Workbook()
    wb.remove(wb.active)
    _add_review_session_sheet(wb, db, reviews)
    filename = f"sesion_{user.role}_{_export_timestamp()}.xlsx"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/candidates/{candidate_id}", response_model=schemas.CandidateOut)
def get_candidate(
    candidate_id: int,
    division: Optional[str] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    candidate = db.get(models.LocationCandidate, candidate_id)
    if not candidate:
        raise HTTPException(404, "Candidate not found")
    _require_candidate_visible(db, candidate, user, division)
    return _candidate_out(db, candidate)


@app.get("/candidates/{candidate_id}/commune-locations", response_model=list[dict[str, str]])
def get_candidate_commune_locations(
    candidate_id: int,
    division: Optional[str] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    candidate = db.get(models.LocationCandidate, candidate_id)
    if not candidate:
        raise HTTPException(404, "Candidate not found")
    _require_candidate_visible(db, candidate, user, division)
    return _candidate_commune_locations(db, candidate)


@app.get(
    "/candidates/{candidate_id}/project-variables",
    response_model=schemas.CandidateProjectVariablesOut,
)
def get_candidate_project_variables(
    candidate_id: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    candidate = db.get(models.LocationCandidate, candidate_id)
    if not candidate:
        raise HTTPException(404, "Candidate not found")
    _require_candidate_visible(db, candidate, user)
    _ensure_project_variables_allowed(db, candidate, user)
    return _project_variables_out(candidate.id, candidate.project_variables)


@app.put(
    "/candidates/{candidate_id}/project-variables",
    response_model=schemas.CandidateProjectVariablesOut,
)
def save_candidate_project_variables(
    candidate_id: int,
    payload: schemas.CandidateProjectVariablesIn,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    candidate = db.get(models.LocationCandidate, candidate_id)
    if not candidate:
        raise HTTPException(404, "Candidate not found")
    _require_candidate_visible(db, candidate, user)
    _ensure_project_variables_allowed(db, candidate, user)
    values = _clean_project_variables_payload(payload)

    variables = candidate.project_variables
    if variables is None:
        variables = models.CandidateProjectVariables(candidate_id=candidate.id)
        db.add(variables)

    for key, value in values.items():
        setattr(variables, key, value)
    variables.updated_by_id = user.id
    db.add(
        models.Review(
            candidate_id=candidate.id,
            stage=workflow.role_stage(user.role) or candidate.current_stage,
            reviewer_id=user.id,
            action="variables_save",
            note="Variables actualizadas",
            created_at=datetime.now(timezone.utc),
        )
    )
    db.commit()
    db.refresh(variables)
    return _project_variables_out(candidate.id, variables)


@app.post(
    "/candidates/{candidate_id}/project-variables/email-preview",
    response_model=list[schemas.CandidateProjectEmailPlanOut],
)
def preview_candidate_project_variable_emails(
    candidate_id: int,
    payload: schemas.CandidateProjectVariablesEmailPreviewIn,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    candidate = db.get(models.LocationCandidate, candidate_id)
    if not candidate:
        raise HTTPException(404, "Candidate not found")
    _require_candidate_visible(db, candidate, user)
    _ensure_project_variables_allowed(db, candidate, user)
    values = _clean_project_variables_payload(payload.variables)
    return _project_email_plans(db, candidate, values)


@app.post(
    "/candidates/{candidate_id}/project-variables/email",
    response_model=schemas.CandidateProjectVariablesEmailOut,
)
def email_candidate_project_variables(
    candidate_id: int,
    payload: schemas.CandidateProjectVariablesEmailIn,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    candidate = db.get(models.LocationCandidate, candidate_id)
    if not candidate:
        raise HTTPException(404, "Candidate not found")
    _require_candidate_visible(db, candidate, user)
    _ensure_project_variables_allowed(db, candidate, user)
    values = _clean_project_variables_payload(payload.variables)

    variables = candidate.project_variables
    if variables is None:
        variables = models.CandidateProjectVariables(candidate_id=candidate.id)
        db.add(variables)
    for key, value in values.items():
        setattr(variables, key, value)
    variables.updated_by_id = user.id
    db.flush()

    plans = {plan["plan_id"]: plan for plan in _project_email_plans(db, candidate, values)}
    if not payload.messages:
        raise HTTPException(400, "Debe seleccionar al menos un correo para enviar.")
    sent_messages: list[schemas.CandidateProjectSentEmailOut] = []
    selected_plan_ids: set[str] = set()
    for selection in payload.messages:
        if selection.plan_id in selected_plan_ids:
            raise HTTPException(400, "Cada correo solo puede enviarse una vez.")
        selected_plan_ids.add(selection.plan_id)
        plan = plans.get(selection.plan_id)
        if not plan:
            raise HTTPException(400, "El correo seleccionado no corresponde al flujo del local.")

        recipients = _normalize_project_email_addresses(selection.recipients, "Para")
        cc = _normalize_project_email_addresses(selection.cc, "CC")
        if not recipients:
            raise HTTPException(400, f"Seleccione al menos un destinatario Para en {plan['area']}.")
        sent_messages.append(_send_project_email_plan(candidate, values, plan, recipients, cc))

    db.add(
        models.Review(
            candidate_id=candidate.id,
            stage=workflow.role_stage(user.role) or candidate.current_stage,
            reviewer_id=user.id,
            action="variables_email",
            note="; ".join(
                f"{message.area}: {', '.join(message.recipients)}"
                for message in sent_messages
            ),
            created_at=datetime.now(timezone.utc),
        )
    )
    db.commit()
    return schemas.CandidateProjectVariablesEmailOut(
        sent=True,
        messages=sent_messages,
    )


@app.get("/candidates/{candidate_id}/reviews", response_model=list[schemas.ReviewOut])
def candidate_reviews(
    candidate_id: int,
    division: Optional[str] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    """Full audit trail for a candidate â€” powers the card's review history."""
    candidate = db.get(models.LocationCandidate, candidate_id)
    if not candidate:
        raise HTTPException(404, "Candidate not found")
    _require_candidate_visible(db, candidate, user, division)
    reviews = db.scalars(
        select(models.Review)
        .where(models.Review.candidate_id == candidate_id)
        .order_by(models.Review.created_at, models.Review.id)
    ).all()
    return [_review_out(r) for r in reviews]


@app.post("/candidates/{candidate_id}/status", response_model=schemas.CandidateActionOut)
def update_candidate_status(
    candidate_id: int,
    payload: schemas.CandidateStatusUpdate,
    request: Request,
    sort_by: str = "score",
    sort_dir: str = "desc",
    division: Optional[str] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    candidate = db.get(models.LocationCandidate, candidate_id)
    if not candidate:
        raise HTTPException(404, "Candidate not found")
    _require_candidate_visible(db, candidate, user, division)
    if payload.group == "pending":
        current_group = workflow.candidate_group(db, candidate)
        is_approver_return = (
            user.role in workflow.APPROVER_ROLES
            and current_group in {"proposed", "rejected"}
        )
        if user.role != workflow.SYSADMIN and not is_approver_return:
            raise HTTPException(
                403,
                "Only Sysadmin, Arriendos y Patentes, or Gerente from Propuestos or Rechazados can return candidates to Pendientes.",
            )
        if is_approver_return and not (payload.note or "").strip():
            raise HTTPException(
                400,
                "Arriendos y Patentes or Gerente must provide a comment when returning a candidate to Pendientes.",
            )
        stage = (
            user.role
            if is_approver_return
            else candidate.current_stage if candidate.current_stage in workflow.STAGES else workflow.COMITE
        )
        action = "send_back" if is_approver_return else "reopen"
        review = models.Review(
            candidate_id=candidate.id,
            stage=stage,
            reviewer_id=user.id,
            action=action,
            note=(payload.note or "").strip() or None,
            created_at=datetime.now(timezone.utc),
        )
        db.add(review)
        candidate.current_stage = stage
        candidate.status = workflow.RETURNED
        candidate.workflow_group = workflow.PENDING
        candidate.last_action = action
        candidate.last_action_at = review.created_at
        candidate.last_actor_role = user.role
        if is_approver_return:
            candidate.returned_at = candidate.last_action_at
        else:
            candidate.reopened_at = candidate.last_action_at
        db.commit()
    else:
        action = {
            "proposed": "accept",
            "approved": "project",
            "rejected": "reject",
            "study": "study",
            "opening": "opening",
            "skip": "skip",
        }[payload.group]
        _require_current_approval_division(db, candidate, action, payload.note)
        if action == "opening":
            _ensure_franchise_activation_variables(db, candidate)
        if user.role in workflow.COMITE_LIKE_ROLES and action in {"project", "reject"}:
            _ensure_review_session_started(request)
        try:
            workflow.submit_review(db, candidate, user, action, payload.note)
        except workflow.WorkflowError as exc:
            raise HTTPException(409, str(exc))
        db.commit()
    db.refresh(candidate)
    return _action_out(db, candidate, user, sort_by=sort_by, sort_dir=sort_dir, commercial_division=division)


@app.post("/candidates/{candidate_id}/comment", response_model=schemas.ReviewOut)
def comment_candidate(
    candidate_id: int,
    payload: schemas.NoteIn,
    division: Optional[str] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    candidate = db.get(models.LocationCandidate, candidate_id)
    if not candidate:
        raise HTTPException(404, "Candidate not found")
    _require_candidate_visible(db, candidate, user, division)
    note = (payload.note or "").strip()
    if not note:
        raise HTTPException(400, "A comment is required.")
    review = models.Review(
        candidate_id=candidate.id,
        stage=workflow.role_stage(user.role) or candidate.current_stage or workflow.COMITE,
        reviewer_id=user.id,
        action="comment",
        note=note,
        created_at=datetime.now(timezone.utc),
    )
    db.add(review)
    db.commit()
    db.refresh(review)
    return _review_out(review)


@app.post("/candidates/{candidate_id}/review", response_model=schemas.CandidateActionOut)
def review_candidate(
    candidate_id: int,
    payload: schemas.ReviewCreate,
    request: Request,
    sort_by: str = "score",
    sort_dir: str = "desc",
    division: Optional[str] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    candidate = db.get(models.LocationCandidate, candidate_id)
    if not candidate:
        raise HTTPException(404, "Candidate not found")
    _require_candidate_visible(db, candidate, user, division)
    _require_current_approval_division(db, candidate, payload.action, payload.note)
    if user.role in workflow.COMITE_LIKE_ROLES and payload.action in {"accept", "project", "reject"}:
        _ensure_review_session_started(request)
    try:
        workflow.submit_review(db, candidate, user, payload.action, payload.note)
    except workflow.WorkflowError as exc:
        raise HTTPException(409, str(exc))
    db.commit()
    db.refresh(candidate)
    return _action_out(db, candidate, user, sort_by=sort_by, sort_dir=sort_dir, commercial_division=division)


@app.post("/candidates/{candidate_id}/send-back", response_model=schemas.CandidateOut)
def send_back_candidate(
    candidate_id: int,
    payload: schemas.NoteIn,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    candidate = db.get(models.LocationCandidate, candidate_id)
    if not candidate:
        raise HTTPException(404, "Candidate not found")
    try:
        workflow.send_back(db, candidate, user, payload.note)
    except workflow.WorkflowError as exc:
        raise HTTPException(409, str(exc))
    db.commit()
    db.refresh(candidate)
    return _candidate_out(db, candidate)


@app.post("/candidates/{candidate_id}/reopen", response_model=schemas.CandidateOut)
def reopen_candidate(
    candidate_id: int,
    payload: schemas.NoteIn,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    candidate = db.get(models.LocationCandidate, candidate_id)
    if not candidate:
        raise HTTPException(404, "Candidate not found")
    try:
        workflow.reopen(db, candidate, user, payload.note)
    except workflow.WorkflowError as exc:
        raise HTTPException(409, str(exc))
    db.commit()
    db.refresh(candidate)
    return _candidate_out(db, candidate)


# --------------------------------------------------------------------------- #
# Pipeline overview (sysadmin dashboard)
# --------------------------------------------------------------------------- #
@app.get("/stats")
def stats(
    project_id: Optional[str] = None,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_role("sysadmin")),
):
    """Counts of candidates per active stage and per terminal status."""
    return _stats_payload(db, project_id)


# --------------------------------------------------------------------------- #
# User management (sysadmin)
# --------------------------------------------------------------------------- #
@app.post("/users", response_model=schemas.UserOut)
def create_user(
    payload: schemas.UserCreate,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_role("sysadmin")),
):
    if db.scalar(select(models.User).where(models.User.email == payload.email)):
        raise HTTPException(409, "A user with that email already exists")
    commercial_division: Optional[str] = None
    if payload.role == workflow.JEFE_COMERCIAL:
        commercial_division = _normal_commercial_division(payload.commercial_division)
        if not commercial_division:
            raise HTTPException(400, "JefeComercial requires commercial_division: SUCURSAL or FRANQUICIA.")
    elif payload.role == workflow.COORDINADOR:
        commercial_division = _normal_commercial_division(payload.commercial_division)
        if not commercial_division:
            raise HTTPException(400, "Coordinador requires commercial_division: SUCURSAL or FRANQUICIA.")
    elif payload.role == workflow.JEFATURA:
        commercial_division = _normal_jefatura_group(payload.commercial_division)
        if not commercial_division:
            raise HTTPException(400, "Jefatura requires commercial_division: SUCURSAL, FRANQUICIA or APERTURA.")
    user = models.User(
        email=payload.email,
        name=payload.name,
        password_hash=auth.hash_password(payload.password),
        role=payload.role,
        commercial_division=commercial_division,
        job_title=(payload.job_title or "").strip() or None,
        supervisor_emails=(payload.supervisor_emails or "").strip() or None,
        org_x=payload.org_x,
        org_y=payload.org_y,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


@app.get("/users", response_model=list[schemas.UserOut])
def list_users(
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_role("sysadmin")),
):
    return db.scalars(
        select(models.User)
        .where(models.User.deleted_at.is_(None))
        .order_by(models.User.created_at)
    ).all()


@app.put("/users/{user_id}", response_model=schemas.UserOut)
def update_user(
    user_id: str,
    payload: schemas.UserUpdate,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_role("sysadmin")),
):
    user = db.get(models.User, user_id)
    if not user or user.deleted_at is not None:
        raise HTTPException(404, "User not found")

    values = payload.model_dump(exclude_unset=True)
    if "name" in values:
        name = (values.get("name") or "").strip()
        if not name:
            raise HTTPException(400, "Name is required.")
        user.name = name
    if "password" in values:
        password = values.get("password") or ""
        if password:
            user.password_hash = auth.hash_password(password)
    if "role" in values:
        user.role = values["role"]
    if "job_title" in values:
        user.job_title = (values.get("job_title") or "").strip() or None
    if "supervisor_emails" in values:
        user.supervisor_emails = (values.get("supervisor_emails") or "").strip() or None
    if "org_x" in values:
        user.org_x = values.get("org_x")
    if "org_y" in values:
        user.org_y = values.get("org_y")
    if "active" in values and values["active"] is not None:
        user.active = bool(values["active"])

    if user.role == workflow.JEFE_COMERCIAL:
        commercial_division = _normal_commercial_division(values.get("commercial_division") or user.commercial_division)
        if not commercial_division:
            raise HTTPException(400, "JefeComercial requires commercial_division: SUCURSAL or FRANQUICIA.")
        user.commercial_division = commercial_division
    elif user.role == workflow.COORDINADOR:
        commercial_division = _normal_commercial_division(values.get("commercial_division") or user.commercial_division)
        if not commercial_division:
            raise HTTPException(400, "Coordinador requires commercial_division: SUCURSAL or FRANQUICIA.")
        user.commercial_division = commercial_division
    elif user.role == workflow.JEFATURA:
        commercial_division = _normal_jefatura_group(values.get("commercial_division") or user.commercial_division)
        user.commercial_division = commercial_division or "APERTURA"
    else:
        user.commercial_division = None

    db.commit()
    db.refresh(user)
    return user


@app.delete("/users/{user_id}")
def delete_user(
    user_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role("sysadmin")),
):
    if user_id == current_user.id:
        raise HTTPException(400, "You cannot delete your own user.")
    user = db.get(models.User, user_id)
    if not user or user.deleted_at is not None:
        raise HTTPException(404, "User not found")
    original_name = user.name
    user.email = f"deleted-{user.id}@deleted.local"
    user.name = f"{original_name} (eliminado)"
    user.password_hash = auth.hash_password(secrets.token_urlsafe(32))
    user.active = False
    user.deleted_at = datetime.now(timezone.utc)
    user.commercial_division = None
    user.job_title = None
    user.supervisor_emails = None
    user.org_x = None
    user.org_y = None
    db.commit()
    return {"ok": True}


# --------------------------------------------------------------------------- #
# Export (sysadmin oversight) â€” per-candidate workflow state + per-stage verdicts
# --------------------------------------------------------------------------- #
@app.get("/projects/{project_id}/results")
def export_results(
    project_id: str,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_role("sysadmin")),
):
    project = db.get(models.Project, project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    candidates = db.scalars(
        select(models.LocationCandidate)
        .where(models.LocationCandidate.project_id == project_id)
        .order_by(models.LocationCandidate.id)
    ).all()

    # Stable union of display_data keys for the trailing columns.
    display_keys: list[str] = []
    for cand in candidates:
        for k in (cand.display_data or {}):
            if k not in display_keys:
                display_keys.append(k)

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    stage_cols: list[str] = []
    for stage in workflow.STAGES:
        stage_cols += [f"{stage}_action", f"{stage}_note"]
    header = (
        ["candidate_id", "status", "current_stage", "priority", "lat", "lng", "map_ref"]
        + stage_cols
        + display_keys
    )
    writer.writerow(header)

    for cand in candidates:
        row = [
            cand.id,
            cand.status,
            cand.current_stage,
            "yes" if cand.priority else "",
            cand.lat if cand.lat is not None else "",
            cand.lng if cand.lng is not None else "",
            cand.map_ref or "",
        ]
        for stage in workflow.STAGES:
            dec = workflow.current_decision(db, cand.id, stage)
            row += [dec.action if dec else "", (dec.note if dec else "") or ""]
        for k in display_keys:
            row.append((cand.display_data or {}).get(k, ""))
        writer.writerow(row)

    buffer.seek(0)
    filename = f"results_{project.name.replace(' ', '_')}.csv"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --------------------------------------------------------------------------- #
# Business locations (global enrichment layer)
# --------------------------------------------------------------------------- #
@app.get("/business", response_model=list[schemas.BusinessOut])
def list_business(
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.get_current_user),
):
    rows = db.scalars(select(models.BusinessLocation)).all()
    result: list[schemas.BusinessOut] = []
    for row in rows:
        attributes = dict(row.attributes or {})
        if "image_url" in attributes:
            attributes["image_url"] = _versioned_image_url(attributes["image_url"])
        result.append(
            schemas.BusinessOut(
                id=row.id,
                name=row.name,
                lat=row.lat,
                lng=row.lng,
                category=row.category,
                attributes=attributes,
            )
        )
    return result


@app.post("/business/ingest", response_model=schemas.BusinessIngestResult)
async def ingest_business(
    file: UploadFile = File(...),
    replace: bool = Form(True),
    config: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_role("sysadmin")),
):
    """Load/replace the global business locations from a tabular file.

    Expects columns: name, lat, lng, category (+ any extras -> attributes).
    Also accepts a single map column ('maps'/'coordinates'/url) if lat/lng absent.
    """
    cfg = _parse_config(config)
    content = await file.read()
    try:
        df = ingestion.read_table(content, file.filename or "business.csv")
    except Exception as exc:
        raise HTTPException(400, f"Could not read tabular file: {exc}")

    if df.empty:
        raise HTTPException(400, "Source file has no rows")

    cols = {c.lower(): c for c in df.columns}
    lat_col = (
        cols.get("lat") or cols.get("latitude") or cols.get("latitud")
    )
    lng_col = (
        cols.get("lng") or cols.get("lon") or cols.get("longitude") or cols.get("longitud")
    )
    # Name: prefer explicit name/nombre, fall back to Direccion + ID composite.
    name_col = cols.get("name") or cols.get("nombre")
    direccion_col = cols.get("direccion") or cols.get("direcciÃ³n")
    id_col = cols.get("idpuntointeres") or cols.get("id")
    cat_col = cols.get("category") or cols.get("categoria") or cols.get("categorÃ­a")
    # For Ahumada data: use Region as category fallback.
    region_col = cols.get("region") or cols.get("regiÃ³n")

    # Optional map column fallback when lat/lng aren't separate.
    map_col = None
    if not (lat_col and lng_col):
        try:
            map_col = ingestion.resolve_map_column(df, cfg.map_column)
        except ValueError:
            raise HTTPException(
                400, "Provide lat/lng columns or a map/coordinates column"
            )

    if replace:
        db.query(models.BusinessLocation).delete()

    created = 0
    failed = 0
    reserved = {c for c in (lat_col, lng_col, name_col, cat_col, map_col) if c}

    for _, row in df.iterrows():
        lat = lng = None
        if lat_col and lng_col:
            lat = ingestion._parse_coord_float(row.get(lat_col))
            lng = ingestion._parse_coord_float(row.get(lng_col))
        elif map_col:
            lat, lng = ingestion.parse_map_ref(row.get(map_col))

        if lat is None or lng is None or not ingestion._valid(lat, lng):
            failed += 1
            continue

        # Build display name: explicit column > Direccion + ID > None.
        if name_col:
            display_name = str(row[name_col])
        elif direccion_col and id_col:
            display_name = f"{row.get(id_col, '')} â€” {row.get(direccion_col, '')}".strip(" â€”")
        elif direccion_col:
            display_name = str(row[direccion_col])
        else:
            display_name = None

        # Category: explicit > Region fallback.
        if cat_col:
            category = str(row[cat_col])
        elif region_col:
            category = str(row[region_col])
        else:
            category = None

        # Attributes: all non-reserved columns, skipping NULL/empty values.
        attributes = {}
        for c in df.columns:
            if c in reserved:
                continue
            v = row.get(c)
            v_str = "" if v is None else str(v).strip()
            if v_str and v_str.upper() != "NULL":
                attributes[c] = v_str

        db.add(
            models.BusinessLocation(
                name=display_name,
                lat=lat,
                lng=lng,
                category=category,
                attributes=attributes,
            )
        )
        created += 1

    db.commit()
    return schemas.BusinessIngestResult(
        rows_read=len(df), locations_created=created, failed_coordinates=failed
    )


# --------------------------------------------------------------------------- #
# Postgres import (sysadmin-triggered, no startup side effects)
# --------------------------------------------------------------------------- #
def _candidate_source_id(display_data: dict) -> str | None:
    for key, value in (display_data or {}).items():
        normalized_key = str(key).strip().lower()
        is_projection_id = (
            normalized_key == "id"
            or (normalized_key.startswith("id ") and "proyecci" in normalized_key)
            or normalized_key.startswith("id proyeccion")
        )
        if is_projection_id and value is not None and str(value).strip():
            return str(value).strip()
    return None


def _candidate_by_projection_id(
    db: Session,
    projection_id: str,
) -> models.LocationCandidate | None:
    requested = str(projection_id or "").strip()
    if not requested:
        return None
    for candidate in db.scalars(select(models.LocationCandidate).order_by(models.LocationCandidate.id)).all():
        if _candidate_source_id(candidate.display_data or {}) == requested:
            return candidate
    return None


def _candidate_audit_payload(db: Session, candidate: models.LocationCandidate) -> dict:
    reviews = db.scalars(
        select(models.Review)
        .where(models.Review.candidate_id == candidate.id)
        .order_by(models.Review.created_at, models.Review.id)
    ).all()
    return {
        "candidate_id": candidate.id,
        "id_proyeccion": _candidate_source_id(candidate.display_data or {}),
        "estado_actual": candidate.status,
        "grupo_actual": workflow.candidate_group(db, candidate),
        "etapa_actual": candidate.current_stage,
        "ultima_accion": candidate.last_action,
        "ultimo_actor_rol": candidate.last_actor_role,
        "ultima_accion_utc": _as_utc_datetime(candidate.last_action_at).isoformat()
        if _as_utc_datetime(candidate.last_action_at)
        else None,
        "ultima_accion_santiago": _santiago_iso(candidate.last_action_at),
        "fechas_estado": {
            "sugerido_santiago": _santiago_iso(candidate.suggested_at),
            "aprobado_santiago": _santiago_iso(candidate.approved_at),
            "rechazado_santiago": _santiago_iso(candidate.rejected_at),
            "proyecto_santiago": _santiago_iso(candidate.project_at),
            "omitido_santiago": _santiago_iso(candidate.skipped_at),
            "devuelto_santiago": _santiago_iso(candidate.returned_at),
            "reabierto_santiago": _santiago_iso(candidate.reopened_at),
            "rechazado_desde_aprobado_santiago": _santiago_iso(candidate.rejected_from_approved_at),
            "rechazado_desde_proyecto_santiago": _santiago_iso(candidate.rejected_from_project_at),
        },
        "movimientos": [
            {
                "revision_id": review.id,
                "accion": review.action,
                "etapa": review.stage,
                "comentario": review.note,
                "usuario_id": review.reviewer_id,
                "usuario_nombre": review.reviewer.name if review.reviewer else None,
                "usuario_correo": review.reviewer.email if review.reviewer else None,
                "usuario_rol": review.reviewer.role if review.reviewer else None,
                "fecha_utc": _as_utc_datetime(review.created_at).isoformat()
                if _as_utc_datetime(review.created_at)
                else None,
                "fecha_santiago": _santiago_iso(review.created_at),
            }
            for review in reviews
        ],
    }


def _get_or_create_postgres_project(db: Session) -> tuple[models.Project, bool]:
    project = db.scalar(
        select(models.Project)
        .where(models.Project.source_file == "postgres")
        .order_by(models.Project.created_at.asc())
    )
    if project:
        return project, False
    project = models.Project(name=POSTGRES_SYNC_PROJECT_NAME, source_file="postgres")
    db.add(project)
    db.flush()
    return project, True


def _upsert_candidate_records(
    db: Session,
    records: list[dict],
    project_id: str,
    replace: bool = False,
) -> tuple[int, int]:
    if replace:
        existing = db.scalars(
            select(models.LocationCandidate)
            .where(models.LocationCandidate.project_id == project_id)
        ).all()
        for candidate in existing:
            db.delete(candidate)
        db.flush()

    existing_by_source_id: dict[str, models.LocationCandidate] = {}
    for candidate in db.scalars(select(models.LocationCandidate)).all():
        source_id = _candidate_source_id(candidate.display_data or {})
        if source_id:
            existing_by_source_id[source_id] = candidate

    created = 0
    updated = 0
    for rec in records:
        display_data = rec.get("display_data") or {}
        source_group = ingestion.candidate_source_group(display_data)
        source_status = workflow.GROUP_TO_DB[source_group]
        source_rejected_at = _candidate_source_date(display_data) if source_group == "observation" else None
        source_id = _candidate_source_id(display_data)
        candidate = existing_by_source_id.get(source_id) if source_id else None
        if candidate is None:
            db.add(
                models.LocationCandidate(
                    project_id=project_id,
                    map_ref=rec["map_ref"],
                    lat=rec["lat"],
                    lng=rec["lng"],
                    display_data=rec["display_data"],
                    current_stage=workflow.JEFATURA,
                    status=source_status,
                    workflow_group=source_status,
                    rejected_at=source_rejected_at,
                )
            )
            created += 1
        else:
            previous_source_group = ingestion.candidate_source_group(candidate.display_data or {})
            candidate.map_ref = rec["map_ref"]
            candidate.lat = rec["lat"]
            candidate.lng = rec["lng"]
            candidate.display_data = rec["display_data"]
            if source_group == "observation" and not candidate.last_action:
                candidate.status = workflow.OBSERVATION
                candidate.workflow_group = workflow.OBSERVATION
                candidate.rejected_at = source_rejected_at
            elif (
                previous_source_group == "observation"
                and workflow.candidate_group(db, candidate) in {"observation", "rejected"}
                and not candidate.last_action
            ):
                candidate.status = workflow.PENDING
                candidate.workflow_group = workflow.PENDING
                candidate.current_stage = workflow.JEFATURA
                candidate.rejected_at = None
            updated += 1
    return created, updated


def _replace_business_records(db: Session) -> tuple[int, int, int]:
    business_records, failed_business_coordinates = (
        ingestion.fetch_business_records_from_postgres()
    )
    db.query(models.BusinessLocation).delete()
    db.flush()
    for rec in business_records:
        db.add(
            models.BusinessLocation(
                name=rec["name"],
                lat=rec["lat"],
                lng=rec["lng"],
                category=rec["category"],
                attributes=rec["attributes"],
            )
        )
    rows_read = len(business_records) + failed_business_coordinates
    return rows_read, len(business_records), failed_business_coordinates


def _sync_postgres(
    db: Session,
    payload: schemas.PostgresImportRequest,
) -> schemas.PostgresImportResult:
    project: models.Project | None = None
    project_created = False
    candidate_rows_read = 0
    candidates_created = 0
    parsed_candidate_coordinates = 0
    failed_candidate_coordinates = 0
    business_rows_read = 0
    business_locations_created = 0
    failed_business_coordinates = 0

    if payload.import_candidates:
        if payload.project_id:
            project = db.get(models.Project, payload.project_id)
            if not project:
                raise HTTPException(404, "Project not found")
        else:
            project, project_created = _get_or_create_postgres_project(db)
            if payload.project_name:
                project.name = payload.project_name

        records, parsed_candidate_coordinates, failed_candidate_coordinates = (
            ingestion.fetch_candidate_records_from_postgres(project.project_id)
        )
        candidate_rows_read = len(records)
        candidates_created, _ = _upsert_candidate_records(
            db,
            records,
            project.project_id,
            replace=payload.replace_candidates,
        )
        project.source_file = "postgres"

    if payload.import_business:
        if payload.replace_business:
            business_rows_read, business_locations_created, failed_business_coordinates = (
                _replace_business_records(db)
            )
        else:
            business_records, failed_business_coordinates = (
                ingestion.fetch_business_records_from_postgres()
            )
            business_rows_read = len(business_records) + failed_business_coordinates
            for rec in business_records:
                db.add(
                    models.BusinessLocation(
                        name=rec["name"],
                        lat=rec["lat"],
                        lng=rec["lng"],
                        category=rec["category"],
                        attributes=rec["attributes"],
                    )
                )
            business_locations_created = len(business_records)

    db.commit()
    return schemas.PostgresImportResult(
        project_id=project.project_id if project else payload.project_id,
        project_created=project_created,
        candidate_rows_read=candidate_rows_read,
        candidates_created=candidates_created,
        parsed_candidate_coordinates=parsed_candidate_coordinates,
        failed_candidate_coordinates=failed_candidate_coordinates,
        business_rows_read=business_rows_read,
        business_locations_created=business_locations_created,
        failed_business_coordinates=failed_business_coordinates,
        replaced_candidates=payload.replace_candidates,
        replaced_business=payload.replace_business,
    )


async def _postgres_sync_loop() -> None:
    while True:
        await asyncio.sleep(POSTGRES_SYNC_INTERVAL_SECONDS)
        await asyncio.to_thread(_run_postgres_sync_once, "interval")


def _run_postgres_sync_once(reason: str = "manual") -> None:
    if not postgres_sync_lock.acquire(blocking=False):
        logger.info("Postgres sync skipped (%s): another sync is already running", reason)
        return
    db = SessionLocal()
    try:
        result = _sync_postgres(
            db,
            schemas.PostgresImportRequest(
                import_candidates=True,
                import_business=True,
                replace_candidates=False,
                replace_business=True,
            ),
        )
        logger.info(
            "Postgres sync completed (%s): candidates read=%s created=%s business=%s",
            reason,
            result.candidate_rows_read,
            result.candidates_created,
            result.business_locations_created,
        )
    except Exception:
        db.rollback()
        logger.exception("Postgres sync failed (%s)", reason)
    finally:
        db.close()
        postgres_sync_lock.release()


@app.post("/admin/import-postgres", response_model=schemas.PostgresImportResult)
def import_postgres(
    payload: schemas.PostgresImportRequest,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_role("sysadmin")),
):
    """Import candidates and global points of interest from the configured Postgres DB."""
    if not postgres_sync_lock.acquire(blocking=False):
        raise HTTPException(409, "Postgres sync is already running.")
    try:
        return _sync_postgres(db, payload)
    except HTTPException:
        db.rollback()
        raise
    except RuntimeError as exc:
        db.rollback()
        raise HTTPException(400, str(exc))
    except Exception as exc:
        db.rollback()
        raise HTTPException(502, f"Postgres import failed: {exc}")
    finally:
        postgres_sync_lock.release()


# --------------------------------------------------------------------------- #
# Frontend (mounted last so API routes win)
# --------------------------------------------------------------------------- #
_VERSIONED_ASSETS = ("app.js", "onboarding.js", "style.css")


def _asset_version(filename: str) -> str:
    """Cache-busting token derived from the asset's mtime (never hardcoded)."""
    try:
        return str(int((STATIC_DIR / filename).stat().st_mtime))
    except OSError:
        return "0"


def _render_index() -> HTMLResponse:
    """Serve index.html with every static asset URL versioned by file mtime, so
    each deploy automatically invalidates the browser cache without a manual bump."""
    html = (STATIC_DIR / "index.html").read_text(encoding="utf-8")
    for name in _VERSIONED_ASSETS:
        html = re.sub(
            rf"/static/{re.escape(name)}(?:\?v=[^\"'>\s]*)?",
            f"/static/{name}?v={_asset_version(name)}",
            html,
        )
    return HTMLResponse(html)


@app.get("/")
def index():
    return _render_index()


@app.get("/ID={projection_id}")
def index_projection(projection_id: str):
    return _render_index()


app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
if IMAGE_DIR.exists():
    app.mount("/images", StaticFiles(directory=IMAGE_DIR), name="images")

