"""API integration test for the approval and project-role flow."""
import base64
import io
import os
import shutil
import tempfile

db_path = os.path.join(tempfile.gettempdir(), "ss_role_flow_api.db")
attachments_path = os.path.join(tempfile.gettempdir(), "ss_role_flow_attachments")
os.environ.pop("DATABASE_URL", None)
os.environ.pop("SITE_SWIPER_DATABASE_URL", None)
os.environ["SITE_SWIPER_DB"] = db_path
os.environ["POSTGRES_AUTO_SYNC"] = "false"
os.environ["SESSION_SECRET"] = "role-flow-test"
os.environ["SYSADMIN_EMAIL"] = "admin@role-flow.test"
os.environ["SYSADMIN_PASSWORD"] = "admin-password"
os.environ["PROJECTION_DOCUMENTS_DIR"] = attachments_path
if os.path.exists(db_path):
    os.remove(db_path)
if os.path.exists(attachments_path):
    shutil.rmtree(attachments_path)

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app import main as main_module, models, workflow  # noqa: E402
from app.database import SessionLocal  # noqa: E402
from app.main import app  # noqa: E402


approval_notifications = []


class ApprovalNotificationSMTP:
    def __init__(self, *_args, **_kwargs):
        pass

    def __enter__(self):
        return self

    def __exit__(self, *_args):
        return False

    def send_message(self, message, from_addr, to_addrs):
        approval_notifications.append((message, from_addr, to_addrs))


main_module.smtplib.SMTP = ApprovalNotificationSMTP


def login(client: TestClient, email: str, password: str) -> None:
    response = client.post("/auth/login", json={"email": email, "password": password})
    assert response.status_code == 200, response.text


with TestClient(app) as admin:
    login(admin, "admin@role-flow.test", "admin-password")
    user_ids = {}
    users = (
        ("arriendo", workflow.ARRIENDO, None),
        ("gerente", workflow.GERENTE, None),
        ("comite", workflow.COMITE, None),
        ("general", workflow.GERENTE_GENERAL, None),
        ("viewer", workflow.VIEWER_GERENTE, None),
        ("jefatura", workflow.JEFATURA, "APERTURA"),
        ("coordinador", workflow.COORDINADOR, "SUCURSAL"),
        ("jefecomercial", workflow.JEFE_COMERCIAL, "SUCURSAL"),
    )
    for email_prefix, role, division in users:
        payload = {
            "email": f"{email_prefix}@role-flow.test",
            "name": email_prefix.title(),
            "password": "test-password",
            "role": role,
        }
        if division:
            payload["commercial_division"] = division
        if role == workflow.JEFE_COMERCIAL:
            payload["supervisor_emails"] = "supervisor@role-flow.test"
        response = admin.post("/users", json=payload)
        assert response.status_code == 200, response.text
        user_ids[role] = response.json()["id"]

    db = SessionLocal()
    project = models.Project(name="Role flow")
    db.add(project)
    db.flush()
    candidate = models.LocationCandidate(
        project_id=project.project_id,
        display_data={"ID": "FLOW-1", "DIVISION": "SUCURSAL"},
        status=workflow.PENDING,
        workflow_group=workflow.PENDING,
    )
    db.add(candidate)
    own_pending = models.LocationCandidate(
        project_id=project.project_id,
        display_data={
            "ID": "OWN-PENDING",
            "DIVISION": "SUCURSAL",
            "CorreoSolicitante": "jefecomercial@role-flow.test",
        },
        status=workflow.PENDING,
        workflow_group=workflow.PENDING,
    )
    own_proposed = models.LocationCandidate(
        project_id=project.project_id,
        display_data={
            "ID": "OWN-PROPOSED",
            "DIVISION": "SUCURSAL",
            "CorreoSolicitante": "JEFEComercial@role-flow.test",
        },
        status=workflow.APPROVED_FINAL,
        workflow_group=workflow.APPROVED_FINAL,
    )
    admin_approved = models.LocationCandidate(
        project_id=project.project_id,
        display_data={
            "ID": "503",
            "DIVISION": "SUCURSAL",
            "NombreSolicitante": "SOLICITANTE ADMIN",
            "CorreoSolicitante": "jefecomercial@role-flow.test",
            "NomComuna": "SANTIAGO CENTRO",
            "NomRegion": "METROPOLITANA DE SANTIAGO",
            "MT2": "80.5 MT2",
            "ValorArriendo": "53.92 UF",
            "CveUnidadCercana": (
                "F0591, STRIP CENTER LAS PERDICES - 957 mts (PROYECTO), "
                "F0034, LAS PARCELAS - 1516 mts (ABIERTA)"
            ),
        },
        status=workflow.PROJECT,
        workflow_group=workflow.PROJECT,
    )
    arriendo_proposed = models.LocationCandidate(
        project_id=project.project_id,
        display_data={"ID": "ARRIENDO-FLOW", "DIVISION": "SUCURSAL"},
        status=workflow.APPROVED_FINAL,
        workflow_group=workflow.APPROVED_FINAL,
    )
    gerente_proposed = models.LocationCandidate(
        project_id=project.project_id,
        display_data={"ID": "GERENTE-APPROVAL", "DIVISION": "SUCURSAL"},
        status=workflow.APPROVED_FINAL,
        workflow_group=workflow.APPROVED_FINAL,
    )
    study_pending = models.LocationCandidate(
        project_id=project.project_id,
        display_data={"ID": "STUDY-PENDING", "DIVISION": "SUCURSAL"},
        status=workflow.PENDING,
        workflow_group=workflow.PENDING,
    )
    study_observation = models.LocationCandidate(
        project_id=project.project_id,
        display_data={"ID": "STUDY-OBSERVATION", "DIVISION": "SUCURSAL"},
        status=workflow.OBSERVATION,
        workflow_group=workflow.OBSERVATION,
    )
    viewer_study = models.LocationCandidate(
        project_id=project.project_id,
        display_data={
            "ID": "VIEW-STUDY",
            "DIVISION": "FRANQUICIA",
            "CorreoSolicitante": "franfrancisco@porunpaismejor.com.mx",
        },
        status=workflow.STUDY,
        workflow_group=workflow.STUDY,
    )
    viewer_proposed = models.LocationCandidate(
        project_id=project.project_id,
        display_data={
            "ID": "VIEW-PROPOSED",
            "DIVISION": "SUCURSAL",
            "CorreoSolicitante": "admjennifer@porunpaismejor.com.mx",
            "NomComuna": "PROVIDENCIA",
            "NomRegion": "METROPOLITANA DE SANTIAGO",
            "MT2": 72,
            "ValorArriendo": "48 UF",
        },
        status=workflow.APPROVED_FINAL,
        workflow_group=workflow.APPROVED_FINAL,
    )
    viewer_approved = models.LocationCandidate(
        project_id=project.project_id,
        display_data={
            "ID": "VIEW-APPROVED",
            "CorreoSolicitante": "aypcelia@porunpaismejor.com.mx",
        },
        status=workflow.PROJECT,
        workflow_group=workflow.PROJECT,
    )
    viewer_opening = models.LocationCandidate(
        project_id=project.project_id,
        display_data={"ID": "VIEW-OPENING", "DIVISION": "FRANQUICIA"},
        status=workflow.OPENING,
        workflow_group=workflow.OPENING,
    )
    viewer_rejected = models.LocationCandidate(
        project_id=project.project_id,
        display_data={"ID": "VIEW-REJECTED", "DIVISION": "SUCURSAL"},
        status=workflow.REJECTED,
        workflow_group=workflow.REJECTED,
    )
    attachment_proposed = models.LocationCandidate(
        project_id=project.project_id,
        display_data={"ID": "501", "DIVISION": "SUCURSAL"},
        status=workflow.APPROVED_FINAL,
        workflow_group=workflow.APPROVED_FINAL,
    )
    attachment_pending = models.LocationCandidate(
        project_id=project.project_id,
        display_data={"ID": "502", "DIVISION": "SUCURSAL"},
        status=workflow.PENDING,
        workflow_group=workflow.PENDING,
    )
    db.add_all([
        own_pending,
        own_proposed,
        admin_approved,
        arriendo_proposed,
        gerente_proposed,
        study_pending,
        study_observation,
        viewer_study,
        viewer_proposed,
        viewer_approved,
        viewer_opening,
        viewer_rejected,
        attachment_proposed,
        attachment_pending,
    ])
    db.commit()
    candidate_id = candidate.id
    own_pending_id = own_pending.id
    own_proposed_id = own_proposed.id
    admin_approved_id = admin_approved.id
    arriendo_proposed_id = arriendo_proposed.id
    gerente_proposed_id = gerente_proposed.id
    study_pending_id = study_pending.id
    study_observation_id = study_observation.id
    viewer_study_id = viewer_study.id
    viewer_proposed_id = viewer_proposed.id
    viewer_approved_id = viewer_approved.id
    viewer_opening_id = viewer_opening.id
    viewer_rejected_id = viewer_rejected.id
    attachment_proposed_id = attachment_proposed.id
    attachment_pending_id = attachment_pending.id
    db.close()

arriendo = TestClient(app)
gerente = TestClient(app)
comite = TestClient(app)
general = TestClient(app)
viewer = TestClient(app)
jefatura = TestClient(app)
coordinador = TestClient(app)
jefe_comercial = TestClient(app)
login(arriendo, "arriendo@role-flow.test", "test-password")
login(gerente, "gerente@role-flow.test", "test-password")
login(comite, "comite@role-flow.test", "test-password")
login(general, "general@role-flow.test", "test-password")
login(viewer, "viewer@role-flow.test", "test-password")
login(jefatura, "jefatura@role-flow.test", "test-password")
login(coordinador, "coordinador@role-flow.test", "test-password")
login(jefe_comercial, "jefecomercial@role-flow.test", "test-password")

# Sysadmin can create the ViewerGerente role through the public user API.
response = admin.post(
    "/users",
    json={
        "name": "Viewer API",
        "email": "viewer-api@role-flow.test",
        "password": "test-password",
        "role": workflow.VIEWER_GERENTE,
    },
)
assert response.status_code == 200, response.text
assert response.json()["role"] == workflow.VIEWER_GERENTE

# ViewerGerente sees every division only in the four read-only groups.
response = viewer.get("/candidates")
assert response.status_code == 200, response.text
viewer_candidates = response.json()
viewer_groups = {item["workflow_group"] for item in viewer_candidates}
assert viewer_groups == {"study", "proposed", "approved", "opening"}, viewer_groups
viewer_ids = {item["id"] for item in viewer_candidates}
assert {
    viewer_study_id,
    viewer_proposed_id,
    viewer_approved_id,
    viewer_opening_id,
}.issubset(viewer_ids)
viewer_categories = {
    item["requested_by"]
    for item in viewer_candidates
    if item["id"] in {viewer_study_id, viewer_proposed_id, viewer_approved_id}
}
assert viewer_categories == {"Sucursal", "Franquicia", "Arriendos"}, viewer_categories
assert candidate_id not in viewer_ids
assert study_observation_id not in viewer_ids
assert viewer_rejected_id not in viewer_ids
assert viewer.get("/candidates/by-projection/VIEW-STUDY").status_code == 200
assert viewer.get("/candidates/by-projection/FLOW-1").status_code == 403

response = viewer.get("/candidates/export.xlsx?group=proposed")
assert response.status_code == 200, response.text
assert response.headers["content-type"].startswith(
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
assert viewer.get("/candidates/export.xlsx?group=study").status_code == 403
assert viewer.get("/candidates/export.xlsx?all_groups=true").status_code == 403
response = gerente.get("/candidates/export.xlsx?all_groups=true")
assert response.status_code == 200, response.text
export_book = load_workbook(io.BytesIO(response.content), read_only=True)
assert export_book.sheetnames == ["Todos los locales"]
export_sheet = export_book["Todos los locales"]
export_header = [cell.value for cell in next(export_sheet.iter_rows(min_row=1, max_row=1))]
group_column = export_header.index("grupo") + 1
export_groups = [
    row[0]
    for row in export_sheet.iter_rows(
        min_row=2,
        min_col=group_column,
        max_col=group_column,
        values_only=True,
    )
    if row[0]
]
export_order = {label: index for index, label in enumerate(main_module.EXPORT_GROUPS.values())}
assert [export_order[group] for group in export_groups] == sorted(export_order[group] for group in export_groups)
export_book.close()
assert viewer.post(
    f"/candidates/{own_proposed_id}/comment",
    json={"note": "No debe guardarse"},
).status_code == 403
assert viewer.post(
    f"/candidates/{viewer_study_id}/status",
    json={"group": "proposed", "note": "No debe avanzar"},
).status_code == 403
assert viewer.post(
    f"/candidates/{own_proposed_id}/review",
    json={"action": "reject", "note": "No debe rechazarse"},
).status_code == 403
assert viewer.get(f"/candidates/{own_proposed_id}/attachments").status_code == 403
assert viewer.get(f"/candidates/{viewer_approved_id}/project-variables").status_code == 403
response = viewer.get(f"/candidates/{viewer_proposed_id}/project-sheet.pdf")
assert response.status_code == 200, response.text
assert response.content.startswith(b"%PDF-")
response = viewer.get(f"/candidates/{viewer_approved_id}/project-sheet.pdf")
assert response.status_code == 200, response.text
assert response.content.startswith(b"%PDF-")
response = viewer.get(f"/candidates/{viewer_opening_id}/project-sheet.pdf")
assert response.status_code == 409, response.text
assert "CveUnidad, Unidad" in response.json()["detail"]

# Gerente can approve a Propuesto and sends the division notification.
notifications_before = len(approval_notifications)
response = gerente.post(
    f"/candidates/{gerente_proposed_id}/status",
    json={"group": "approved", "note": "División: FRANQUICIA"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "approved"
assert len(approval_notifications) == notifications_before + 1
approval_message, approval_from, approval_to = approval_notifications[-1]
assert approval_from == "mbustos@farmaciasdoctorsimi.cl"
assert approval_to == [
    "dcastro@farmaciasdoctorsimi.cl",
    "mcasanova@porunpaismejor.com.mx",
    "admjennifer@porunpaismejor.com.mx",
    "lalbornoz@farmaciasdoctorsimi.cl",
    "mbustos@farmaciasdoctorsimi.cl",
    "rmalave@farmaciasdoctorsimi.cl",
]
assert approval_message["From"] == "mbustos@farmaciasdoctorsimi.cl"
assert approval_message["To"] == (
    "dcastro@farmaciasdoctorsimi.cl, mcasanova@porunpaismejor.com.mx, "
    "admjennifer@porunpaismejor.com.mx, lalbornoz@farmaciasdoctorsimi.cl"
)
assert approval_message["Cc"] == (
    "mbustos@farmaciasdoctorsimi.cl, rmalave@farmaciasdoctorsimi.cl"
)
approval_body = approval_message.get_body(preferencelist=("plain",)).get_content()
approval_html = approval_message.get_body(preferencelist=("html",)).get_content()
assert approval_message["Subject"] == "Proyección aprobada | ID GERENTE-APPROVAL"
assert "ID: GERENTE-APPROVAL" in approval_body
assert "Destino: FRANQUICIA" in approval_body
assert "http://172.23.1.128:8002/ID=GERENTE-APPROVAL" in approval_body
assert "Mauricio Bustos Miranda" in approval_body
assert "Analista de datos" in approval_body
assert "Revisar proyección" in approval_html
assert "Mauricio Bustos Miranda" in approval_html

# Projection deep links resolve visible candidates in any workflow group.
response = gerente.get("/candidates/by-projection/OWN-PENDING")
assert response.status_code == 200, response.text
assert response.json()["workflow_group"] == "pending"
response = gerente.get("/candidates/by-projection/OWN-PROPOSED")
assert response.status_code == 200, response.text
assert response.json()["workflow_group"] == "proposed"
response = gerente.get("/candidates/by-projection/503")
assert response.status_code == 200, response.text
assert response.json()["workflow_group"] == "approved"
response = gerente.get("/candidates/by-projection/STUDY-OBSERVATION")
assert response.status_code == 200, response.text
assert response.json()["workflow_group"] == "observation"

# Projection files are stored by numeric projection and remain readable later.
png_bytes = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
)
response = gerente.post(
    f"/candidates/{attachment_pending_id}/attachments",
    files={"files": ("pendiente.png", png_bytes, "image/png")},
)
assert response.status_code == 409, response.text
response = gerente.post(
    f"/candidates/{attachment_proposed_id}/attachments",
    files={"files": ("frontis.png", png_bytes, "image/png")},
)
assert response.status_code == 200, response.text
attachments = response.json()
assert len(attachments) == 1
assert attachments[0]["name"] == "frontis.png"
assert os.path.exists(os.path.join(attachments_path, "Proyeccion501", "frontis.png"))
response = gerente.get(f"/candidates/{attachment_proposed_id}/attachments")
assert response.status_code == 200, response.text
assert response.json()[0]["content_type"] == "image/png"
response = gerente.get(response.json()[0]["url"])
assert response.status_code == 200, response.text
assert response.content == png_bytes
response = gerente.post(
    f"/candidates/{attachment_proposed_id}/attachments",
    files=[
        ("files", ("contrato.pdf", b"%PDF-1.4\n%%EOF", "application/pdf")),
        (
            "files",
            (
                "presentacion.pptx",
                b"PK\x03\x04office-content",
                "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            ),
        ),
    ],
)
assert response.status_code == 200, response.text
assert {item["name"] for item in response.json()} == {
    "frontis.png",
    "contrato.pdf",
    "presentacion.pptx",
}
pdf_attachment = next(item for item in response.json() if item["name"] == "contrato.pdf")
response = gerente.get(pdf_attachment["url"])
assert response.status_code == 200, response.text
assert response.headers["content-type"].startswith("application/pdf")
assert response.headers["content-disposition"].startswith("inline;")
response = gerente.delete(
    f"/candidates/{attachment_proposed_id}/attachments/frontis.png",
)
assert response.status_code == 200, response.text
assert {item["name"] for item in response.json()} == {"contrato.pdf", "presentacion.pptx"}
assert not os.path.exists(os.path.join(attachments_path, "Proyeccion501", "frontis.png"))
response = gerente.post(
    f"/candidates/{attachment_proposed_id}/attachments",
    files={"files": ("programa.exe", b"MZ", "application/octet-stream")},
)
assert response.status_code == 400, response.text
response = gerente.post(
    f"/candidates/{attachment_proposed_id}/status",
    json={"group": "rejected", "note": "Validar persistencia de imagen"},
)
assert response.status_code == 200, response.text
response = gerente.get(f"/candidates/{attachment_proposed_id}/attachments")
assert response.status_code == 200, response.text
assert {item["name"] for item in response.json()} == {"contrato.pdf", "presentacion.pptx"}
response = gerente.delete(
    f"/candidates/{attachment_proposed_id}/attachments/contrato.pdf",
)
assert response.status_code == 200, response.text
assert [item["name"] for item in response.json()] == ["presentacion.pptx"]
assert not os.path.exists(os.path.join(attachments_path, "Proyeccion501", "contrato.pdf"))
db = SessionLocal()
attachment_review = db.query(models.Review).filter(
    models.Review.candidate_id == attachment_proposed_id,
    models.Review.action == "attachment_upload",
    models.Review.note == "frontis.png",
).one()
assert attachment_review.note == "frontis.png"
attachment_delete_review = db.query(models.Review).filter(
    models.Review.candidate_id == attachment_proposed_id,
    models.Review.action == "attachment_delete",
    models.Review.note == "frontis.png",
).one()
assert attachment_delete_review.note == "frontis.png"
db.close()

# Comments can be saved without changing the candidate state.
response = gerente.post(
    f"/candidates/{study_pending_id}/comment",
    json={"note": "Comentario independiente de gerencia"},
)
assert response.status_code == 200, response.text
assert response.json()["action"] == "comment"
assert response.json()["note"] == "Comentario independiente de gerencia"
db = SessionLocal()
commented_candidate = db.get(models.LocationCandidate, study_pending_id)
assert workflow.candidate_group(db, commented_candidate) == "pending"
db.close()

# En Estudio accepts Pendientes/Observación and exits to Propuestos/Rechazados.
response = gerente.post(
    f"/candidates/{study_pending_id}/status",
    json={"group": "study", "note": "Local llamativo"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "study"
response = gerente.post(
    f"/candidates/{study_pending_id}/status",
    json={"group": "proposed", "note": "Estudio favorable"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "proposed"

response = gerente.post(
    f"/candidates/{study_observation_id}/status",
    json={"group": "study", "note": "Revisar potencial"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "study"
response = arriendo.post(
    f"/candidates/{study_observation_id}/status",
    json={"group": "rejected", "note": "Estudio descartado"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "rejected"
response = coordinador.post(
    f"/candidates/{own_pending_id}/status",
    json={"group": "study", "note": "Sin permiso"},
)
assert response.status_code == 409, response.text

# Arriendos y Patentes can approve, configure, activate, email, and deactivate.
sync_version_before = arriendo.get("/sync/version")
assert sync_version_before.status_code == 200, sync_version_before.text
response = arriendo.post(
    f"/candidates/{arriendo_proposed_id}/status",
    json={"group": "approved", "note": "Aprobado por Arriendos\nDivisión: SUCURSAL"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "approved"
sync_version_after = arriendo.get("/sync/version")
assert sync_version_after.status_code == 200, sync_version_after.text
assert sync_version_after.json()["version"] != sync_version_before.json()["version"]

# The frequent background cycle imports candidates without reloading heavy business layers.
captured_sync_payloads = []
original_sync_postgres = main_module._sync_postgres


def capture_candidate_sync(_db, payload):
    captured_sync_payloads.append(payload)
    return main_module.schemas.PostgresImportResult(
        project_created=False,
        candidate_rows_read=1,
        candidates_created=1,
        parsed_candidate_coordinates=1,
        failed_candidate_coordinates=0,
        business_rows_read=0,
        business_locations_created=0,
        failed_business_coordinates=0,
        replaced_candidates=False,
        replaced_business=False,
    )


try:
    main_module._sync_postgres = capture_candidate_sync
    main_module._run_postgres_sync_once("test_candidate_interval", import_business=False)
finally:
    main_module._sync_postgres = original_sync_postgres

assert len(captured_sync_payloads) == 1
assert captured_sync_payloads[0].import_candidates is True
assert captured_sync_payloads[0].import_business is False
assert captured_sync_payloads[0].replace_business is False

arriendo_variables = {
    "cve_unidad": "CLAYP",
    "unidad": "LOCAL ARRIENDOS",
    "region": "METROPOLITANA DE SANTIAGO",
    "comuna": "SANTIAGO",
    "contacto_nombre": "CONTACTO ARRIENDOS",
    "contacto_telefono": "+56933333333",
    "contacto_email": "contacto.arriendos@example.com",
}
response = arriendo.put(
    f"/candidates/{arriendo_proposed_id}/project-variables",
    json=arriendo_variables,
)
assert response.status_code == 200, response.text
response = arriendo.post(
    f"/candidates/{arriendo_proposed_id}/status",
    json={"group": "opening", "note": "Alta por Arriendos"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "opening"
response = arriendo.post(
    f"/candidates/{arriendo_proposed_id}/project-variables/email-preview",
    json={"variables": arriendo_variables},
)
assert response.status_code == 200, response.text
assert len(response.json()) == 2
response = arriendo.post(
    f"/candidates/{arriendo_proposed_id}/status",
    json={"group": "rejected", "note": "Baja por Arriendos"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "rejected"

# Sysadmin can perform the Coordinator variable and activation workflow.
admin_actions = TestClient(app)
login(admin_actions, "admin@role-flow.test", "admin-password")
assert main_module._project_sheet_text(None) == ""
assert main_module._project_sheet_nearby_units({
    "CveUnidadCercana": (
        "F0591, STRIP CENTER LAS PERDICES - 957 mts (PROYECTO), "
        "F0034, LAS PARCELAS - 1516 mts (ABIERTA), F0591"
    ),
    "CveUnidadPropiaCercana": "F0591",
}) == [
    "F0591, STRIP CENTER LAS PERDICES - 957 mts (PROYECTO)",
    "F0034, LAS PARCELAS - 1516 mts (ABIERTA)",
]
assert main_module._project_sheet_projection_value(60) == "$60 MM"
response = admin_actions.put(
    f"/candidates/{admin_approved_id}/project-variables",
    json={
        "cve_unidad": "CLADMIN",
        "unidad": "LOCAL ADMIN",
        "proyeccion_supervisor": 65.5,
    },
)
assert response.status_code == 422, response.text
response = admin_actions.get(f"/candidates/{admin_approved_id}/project-variables")
assert response.status_code == 200, response.text
assert response.json()["comuna"] == "SANTIAGO CENTRO"
assert response.json()["region"] == "METROPOLITANA DE SANTIAGO"
assert response.json()["mt2"] == 80.5
assert response.json()["valor_arriendo"] == "53.92 UF"
response = admin_actions.put(
    f"/candidates/{admin_approved_id}/project-variables",
    json={
        "cve_unidad": "CLADMIN",
        "unidad": "LOCAL ADMIN",
        "region": "METROPOLITANA DE SANTIAGO",
        "comuna": "SANTIAGO",
        "tiendas_anclas": "SUPERMERCADO Y ESTACION DE SERVICIO",
        "proyeccion_supervisor": 65,
        "proyeccion_jefe_comercial": 70,
    },
)
assert response.status_code == 200, response.text
assert response.json()["tiendas_anclas"] == "SUPERMERCADO Y ESTACION DE SERVICIO"
assert response.json()["proyeccion_supervisor"] == 65
assert response.json()["proyeccion_jefe_comercial"] == 70
admin_images_path = os.path.join(attachments_path, "Proyeccion503")
os.makedirs(admin_images_path, exist_ok=True)
for image_name in ("fachada.png", "interior.png", "entorno.png"):
    with open(os.path.join(admin_images_path, image_name), "wb") as image_file:
        image_file.write(png_bytes)
db = SessionLocal()
admin_approved_candidate = db.get(models.LocationCandidate, admin_approved_id)
assert len(main_module._project_sheet_photos(admin_approved_candidate)) == 3
db.close()
response = admin_actions.get(f"/candidates/{admin_approved_id}/project-sheet.pdf")
assert response.status_code == 200, response.text
assert response.content.startswith(b"%PDF-")
for role_client in (
    arriendo,
    gerente,
    comite,
    general,
    viewer,
    jefatura,
    coordinador,
    jefe_comercial,
):
    response = role_client.get(f"/candidates/{admin_approved_id}/project-sheet.pdf")
    assert response.status_code == 200, response.text
    assert response.content.startswith(b"%PDF-")
response = admin_actions.post(
    f"/candidates/{admin_approved_id}/status",
    json={"group": "opening", "note": "Alta por sysadmin"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "opening"
response = admin_actions.get(f"/candidates/{admin_approved_id}/project-sheet.pdf")
assert response.status_code == 200, response.text
assert response.headers["content-type"].startswith("application/pdf")
assert "Ficha_Proyecto_503.pdf" in response.headers["content-disposition"]
assert response.content.startswith(b"%PDF-")
assert len(response.content) > 5000
response = coordinador.get(f"/candidates/{admin_approved_id}/project-sheet.pdf")
assert response.status_code == 200, response.text
response = admin_actions.get(f"/candidates/{own_pending_id}/project-sheet.pdf")
assert response.status_code == 409, response.text

# Jefe Comercial can see their own pending/proposed locations, but cannot vote on them.
response = jefe_comercial.get("/candidates")
assert response.status_code == 200, response.text
visible_ids = {item["id"] for item in response.json()}
assert own_pending_id in visible_ids
assert own_proposed_id in visible_ids
response = jefe_comercial.post(
    f"/candidates/{own_pending_id}/review",
    json={"action": "like"},
)
assert response.status_code == 409, response.text

# Arriendo rejects a pending location and Gerente proposes it again.
response = arriendo.post(
    f"/candidates/{own_pending_id}/status",
    json={"group": "rejected", "note": "Antecedentes incompletos"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "rejected"
response = gerente.post(
    f"/candidates/{own_pending_id}/status",
    json={"group": "proposed", "note": "Antecedentes corregidos"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "proposed"
response = gerente.post(
    f"/candidates/{own_pending_id}/status",
    json={"group": "rejected"},
)
assert response.status_code == 409, response.text
response = gerente.post(
    f"/candidates/{own_pending_id}/status",
    json={"group": "rejected", "note": "Gerencia solicita revisar antecedentes"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "rejected"

# Rejected locations can return to Pending, move to En Estudio, or be proposed.
response = gerente.post(
    f"/candidates/{own_pending_id}/status",
    json={"group": "pending"},
)
assert response.status_code == 400, response.text
response = gerente.post(
    f"/candidates/{own_pending_id}/status",
    json={"group": "pending", "note": "Reevaluar desde pendientes"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "pending"
response = arriendo.post(
    f"/candidates/{own_pending_id}/status",
    json={"group": "rejected", "note": "Rechazo para estudio adicional"},
)
assert response.status_code == 200, response.text
response = arriendo.post(
    f"/candidates/{own_pending_id}/status",
    json={"group": "study", "note": "Evaluar alternativa de arriendo"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "study"
response = gerente.post(
    f"/candidates/{own_pending_id}/status",
    json={"group": "rejected", "note": "Alternativa descartada"},
)
assert response.status_code == 200, response.text
response = admin_actions.post(
    f"/candidates/{own_pending_id}/status",
    json={"group": "study", "note": "Revision administrativa"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "study"
response = admin_actions.post(
    f"/candidates/{own_pending_id}/status",
    json={"group": "rejected", "note": "Devuelto a rechazados por admin"},
)
assert response.status_code == 200, response.text
response = gerente.post(
    f"/candidates/{own_pending_id}/status",
    json={"group": "proposed", "note": "Rechazado reconsiderado"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "proposed"

# Gerente General can omit a proposed location without changing its group.
response = general.post(
    f"/candidates/{own_proposed_id}/review",
    json={"action": "skip"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "proposed"

response = gerente.post(
    f"/candidates/{own_proposed_id}/status",
    json={"group": "skip", "note": "Revisión postergada"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "proposed"

# Arriendos y Patentes and Gerente can return a Propuesto to Pendientes with a comment.
response = arriendo.post(
    f"/candidates/{own_proposed_id}/status",
    json={"group": "pending"},
)
assert response.status_code == 400, response.text
response = arriendo.post(
    f"/candidates/{own_proposed_id}/status",
    json={"group": "pending", "note": "Faltan antecedentes de arriendo"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "pending"

response = arriendo.post(
    f"/candidates/{own_proposed_id}/status",
    json={"group": "proposed", "note": "Antecedentes completados"},
)
assert response.status_code == 200, response.text
response = arriendo.post(
    f"/candidates/{own_proposed_id}/status",
    json={"group": "rejected"},
)
assert response.status_code == 409, response.text
response = arriendo.post(
    f"/candidates/{own_proposed_id}/status",
    json={"group": "rejected", "note": "Condiciones de arriendo insuficientes"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "rejected"

response = arriendo.post(
    f"/candidates/{own_proposed_id}/status",
    json={"group": "proposed", "note": "Condiciones corregidas"},
)
assert response.status_code == 200, response.text
response = gerente.post(
    f"/candidates/{own_proposed_id}/status",
    json={"group": "pending"},
)
assert response.status_code == 400, response.text
response = gerente.post(
    f"/candidates/{own_proposed_id}/status",
    json={"group": "pending", "note": "Faltan antecedentes comerciales"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "pending"
db = SessionLocal()
return_review = db.query(models.Review).filter(
    models.Review.candidate_id == own_proposed_id,
    models.Review.action == "send_back",
    models.Review.reviewer_id == user_ids[workflow.GERENTE],
).one()
assert return_review.note == "Faltan antecedentes comerciales"
arriendo_return_review = db.query(models.Review).filter(
    models.Review.candidate_id == own_proposed_id,
    models.Review.action == "send_back",
    models.Review.reviewer_id == user_ids[workflow.ARRIENDO],
).one()
assert arriendo_return_review.note == "Faltan antecedentes de arriendo"
db.close()

response = arriendo.post(f"/candidates/{candidate_id}/status", json={"group": "proposed"})
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "proposed"

response = comite.post(f"/candidates/{candidate_id}/status", json={"group": "approved"})
assert response.status_code == 400, response.text
response = comite.post(
    f"/candidates/{candidate_id}/status",
    json={"group": "approved", "note": "Comentario de aprobación\nDivisión: SUCURSAL"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "approved"
db = SessionLocal()
approval_review = db.query(models.Review).filter(
    models.Review.candidate_id == candidate_id,
    models.Review.action == "project",
).order_by(models.Review.id.desc()).first()
assert approval_review is not None
assert "Comentario de aprobación" in (approval_review.note or "")
assert "División: SUCURSAL" in (approval_review.note or "")
db.close()

assert comite.get(f"/candidates/{candidate_id}/project-variables").status_code == 403
assert coordinador.get(f"/candidates/{candidate_id}/project-variables").status_code == 200

variables = {
    "cve_unidad": "CL9999",
    "unidad": "LOCAL TEST",
    "region": "METROPOLITANA DE SANTIAGO",
    "comuna": "SANTIAGO",
}
response = coordinador.put(f"/candidates/{candidate_id}/project-variables", json=variables)
assert response.status_code == 200, response.text

preview = coordinador.post(
    f"/candidates/{candidate_id}/project-variables/email-preview",
    json={"variables": variables},
)
assert preview.status_code == 200, preview.text
assert [plan["plan_id"] for plan in preview.json()] == ["sucursal_legal", "sucursal_reducido"]
assert all(plan["from_email"] == "admjennifer@porunpaismejor.com.mx" for plan in preview.json())

sent_messages = []


class FakeSMTP:
    def __init__(self, *_args, **_kwargs):
        pass

    def __enter__(self):
        return self

    def __exit__(self, *_args):
        return False

    def send_message(self, message, from_addr, to_addrs):
        sent_messages.append((message, from_addr, to_addrs))


original_smtp = main_module.smtplib.SMTP
main_module.smtplib.SMTP = FakeSMTP
try:
    response = coordinador.post(
        f"/candidates/{candidate_id}/project-variables/email",
        json={
            "variables": variables,
            "messages": [
                {
                    "plan_id": "sucursal_legal",
                    "recipients": ["curibe@farmaciasdoctorsimi.cl"],
                    "cc": ["mcasanova@farmaciasdoctorsimi.cl"],
                }
            ],
        },
    )
finally:
    main_module.smtplib.SMTP = original_smtp
assert response.status_code == 200, response.text
assert response.json()["messages"][0]["plan_id"] == "sucursal_legal"
assert len(sent_messages) == 1
assert sent_messages[0][1] == "admjennifer@porunpaismejor.com.mx"
assert sent_messages[0][2] == [
    "curibe@farmaciasdoctorsimi.cl",
    "mcasanova@farmaciasdoctorsimi.cl",
]

response = coordinador.post(f"/candidates/{candidate_id}/status", json={"group": "opening"})
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "opening"
assert coordinador.get(f"/candidates/{candidate_id}/project-variables").status_code == 200
project_preview = coordinador.post(
    f"/candidates/{candidate_id}/project-variables/email-preview",
    json={"variables": variables},
)
assert project_preview.status_code == 200, project_preview.text
assert [plan["plan_id"] for plan in project_preview.json()] == [
    "sucursal_legal",
    "sucursal_reducido",
]
assert comite.get(f"/candidates/{candidate_id}/project-variables").status_code == 403

response = general.post(
    f"/candidates/{candidate_id}/status",
    json={"group": "rejected", "note": "Dar de baja desde Proyectos"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "rejected"

# Every new Propuesto -> Aprobado cycle requires a fresh division selection.
response = arriendo.post(
    f"/candidates/{candidate_id}/status",
    json={"group": "proposed", "note": "Reevaluar metodología"},
)
assert response.status_code == 200, response.text
response = comite.post(f"/candidates/{candidate_id}/status", json={"group": "approved"})
assert response.status_code == 400, response.text
response = comite.post(
    f"/candidates/{candidate_id}/status",
    json={"group": "approved", "note": "División: FRANQUICIA"},
)
assert response.status_code == 200, response.text
assert response.json()["candidate"]["workflow_group"] == "approved"
assert response.json()["candidate"]["approved_division"] == "FRANQUICIA"

# Sysadmin can delete a Gerente General with history without losing the audit trail.
admin_delete = TestClient(app)
login(admin_delete, "admin@role-flow.test", "admin-password")
general_user_id = user_ids[workflow.GERENTE_GENERAL]
response = admin_delete.delete(f"/users/{general_user_id}")
assert response.status_code == 200, response.text
assert general_user_id not in {user["id"] for user in admin_delete.get("/users").json()}
assert general.post(
    "/auth/login",
    json={"email": "general@role-flow.test", "password": "test-password"},
).status_code == 401

db = SessionLocal()
deleted_general = db.get(models.User, general_user_id)
assert deleted_general is not None
assert deleted_general.deleted_at is not None and deleted_general.active is False
assert db.query(models.Review).filter(models.Review.reviewer_id == general_user_id).count() > 0
db.close()

response = admin_delete.post("/users", json={
    "email": "general@role-flow.test",
    "name": "General Reemplazo",
    "password": "replacement-password",
    "role": workflow.GERENTE_GENERAL,
})
assert response.status_code == 200, response.text

print("ROLE FLOW API TESTS PASSED")
